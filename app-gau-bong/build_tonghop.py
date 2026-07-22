#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""FILE TONG HOP THU CHI - LOI NHUAN 2026 (gop 2 file):
- Moi thang 2 tab cung mau:
  + "Tháng m"  : tong quan (luong 15 dong, mua hang theo tuan, van hanh, doanh thu, LOI NHUAN)
  + "MH Tháng m": chi tiet mua hang theo NGAY (8 dong/ngay, tong ngay/tuan/thang)
- Tab cuoi "TỔNG HỢP NĂM": chi phi + doanh thu + loi nhuan 12 thang + ca nam
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

ARIAL = "Arial"
NAVY = "1F3864"
BLUE_SEC = "2E75B6"
GREEN_SEC = "548235"
ORANGE_SEC = "C55A11"
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
DAY_FILL = PatternFill("solid", fgColor="DDEBF7")
WEEK_FILL = PatternFill("solid", fgColor="FCE4D6")
GRAND_FILL = PatternFill("solid", fgColor="1F3864")
GREEN_FILL = PatternFill("solid", fgColor="375623")
REV_FILL = PatternFill("solid", fgColor="548235")
BAND_FILL = PatternFill("solid", fgColor="EDEDED")
thin = Side(style="thin", color="BFBFBF")
med = Side(style="medium", color="1F3864")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
VND = '#,##0'
TAB_COLORS = ["E53935", "F4511E", "FB8C00", "FDD835", "7CB342", "43A047",
              "00897B", "039BE5", "3949AB", "5E35B1", "8E24AA", "D81B60"]
DAYS = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
N_ROW = 8    # dong hang moi ngay (tab MH)
N_S = 15     # dong tien luong

# layout tab tong quan "Tháng m"
SAL_FIRST, SAL_LAST = 5, 19
ROW_TL = 20
SEC2, HDR2 = 22, 23
WK_FIRST = 24                     # 5 dong tuan: 24-28
ROW_MH = 29
SEC3, HDR3 = 31, 32
ROW_O1 = 33
ROW_T3 = 40
ROW_G = 42
ROW_DT = 44
ROW_LN = 45

def f(bold=False, size=10, white=False, blue=False, color="000000"):
    c = "FFFFFF" if white else ("0D47A1" if blue else color)
    return Font(name=ARIAL, bold=bold, size=size, color=c)

wb = Workbook()

# ============ HUONG DAN ============
ws = wb.active
ws.title = "HƯỚNG DẪN"
ws.sheet_properties.tabColor = "757575"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 112
guide = [
    ("HƯỚNG DẪN — FILE QUẢN LÝ THU CHI & LỢI NHUẬN 2026 (BẢN GỘP)", 1),
    ("", 0),
    ("CẤU TRÚC FILE: mỗi tháng có 2 tab CÙNG MÀU đứng cạnh nhau:", 2),
    ("   • Tab 'Tháng 1' … 'Tháng 12' — TỔNG QUAN tháng: I. Tiền lương (15 dòng) · II. Chi phí mua hàng (tự kéo theo tuần) ·", 0),
    ("     III. Vận hành & phát sinh · 💰 Tổng chi phí · 📈 Doanh thu (bạn nhập) · ✅ LỢI NHUẬN (tự tính).", 0),
    ("   • Tab 'MH Tháng 1' … 'MH Tháng 12' — NHẬP MUA HÀNG THEO NGÀY: mỗi ngày 8 dòng, có tổng NGÀY, tổng TUẦN, tổng THÁNG.", 0),
    ("   • Tab cuối 'TỔNG HỢP NĂM' — chi phí, doanh thu, LỢI NHUẬN 12 tháng + CẢ NĂM.", 0),
    ("", 0),
    ("CÁCH NHẬP LIỆU (chỉ nhập ô NỀN VÀNG, chữ xanh):", 2),
    ("   1. Mua hàng hằng ngày → mở tab 'MH Tháng x', nhập vào đúng ngày: Tên sản phẩm, Số lượng, Đơn giá.", 0),
    ("      Thành tiền tự nhân; tổng ngày/tuần/tháng tự cộng và TỰ CHẢY sang tab 'Tháng x' (phần II).", 0),
    ("   2. Đầu/cuối tháng → mở tab 'Tháng x': nhập lương nhân viên (phần I), tiền điện nước mạng quảng cáo (phần III),", 0),
    ("      và DOANH THU THÁNG ở ô vàng cuối trang.", 0),
    ("   3. Xong! LỢI NHUẬN tháng và cả năm tự tính, xem nhanh ở tab 'TỔNG HỢP NĂM'.", 0),
    ("", 0),
    ("CÔNG THỨC CHÍNH (tự động — không sửa tay):", 2),
    ("   - Thành tiền = Số lượng × Đơn giá        - Tổng tuần: T1=ngày 1–7 · T2=8–14 · T3=15–21 · T4=22–28 · T5=29–hết", 0),
    ("   - TỔNG CHI PHÍ THÁNG = Lương (1) + Mua hàng (2) + Vận hành (3)        - LỢI NHUẬN = Doanh thu − Tổng chi phí", 0),
    ("", 0),
    ("   • Tháng 1 có SỐ VÍ DỤ (chữ xanh) — thay bằng số thật. Tháng 2 chỉ 28 ngày nên chỉ có 4 tuần, ô Tuần 5 trống là đúng.", 0),
    ("   • Ngày nào không mua hàng để trống, tổng tự = 0. Đơn vị: VNĐ.", 0),
]
for i, (text, style) in enumerate(guide, start=1):
    c = ws.cell(row=i, column=2, value=text)
    c.font = Font(name=ARIAL, bold=True, size=14, color=NAVY) if style == 1 else f(bold=(style > 0), size=11)

EX_SALARY = [(2, 9000000), (3, 8000000), (2, 7000000), (4, 8500000)]
EX_OPS = [3500000, 500000, 400000, 15000000, 1200000, None, None]
STAFF = ["Nhân viên marketing", "Nhân viên sale", "Nhân viên đóng hàng", "Nhân viên sản xuất"]
OPS = ["Tiền điện", "Tiền nước", "Tiền mạng (internet)", "Tiền chạy quảng cáo",
       "Chi phí phát sinh 1", "Chi phí phát sinh 2", "Chi phí phát sinh 3"]
EX = {
    (1, 0): ("Hộp quà tặng (ví dụ)", 200, 25000),
    (1, 1): ("Túi giấy in logo (ví dụ)", 500, 8000),
    (1, 2): ("Ruy băng (ví dụ)", 20, 35000),
    (2, 0): ("Bình giữ nhiệt in logo (ví dụ)", 100, 95000),
}

# ============ 12 TAB "MH Tháng m" (mua hang theo ngay) ============
mh_refs = {}   # m -> (week_rows, month_row)
for idx in range(12):
    m = idx + 1
    ndays = DAYS[m]
    ws = wb.create_sheet(f"MH Tháng {m}")
    ws.sheet_properties.tabColor = TAB_COLORS[idx]
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 11, "B": 5, "C": 34, "D": 11, "E": 14, "F": 17}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"MUA HÀNG THEO NGÀY — THÁNG {m}/2026   (số liệu tự chảy sang tab 'Tháng {m}')"
    c.font = f(bold=True, size=13, white=True)
    c.fill = PatternFill("solid", fgColor=TAB_COLORS[idx])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for j, lab in enumerate(["Ngày", "STT", "Tên sản phẩm", "Số lượng", "Đơn giá (VNĐ)", "Thành tiền (VNĐ)"], start=1):
        cc = ws.cell(row=3, column=j, value=lab)
        cc.font = f(bold=True, white=True, size=9)
        cc.fill = PatternFill("solid", fgColor=NAVY)
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cc.border = BORDER
    ws.row_dimensions[3].height = 20

    r = 4
    week_rows = []
    week_day_rows = []
    week_first_day = 1
    for d in range(1, ndays + 1):
        first = r
        ws.merge_cells(f"A{first}:A{first + N_ROW - 1}")
        ac = ws.cell(row=first, column=1, value=f"Ngày {d}/{m}")
        ac.font = f(bold=True, size=9)
        ac.alignment = Alignment(horizontal="center", vertical="center")
        ac.fill = BAND_FILL if d % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for i in range(N_ROW):
            rr = first + i
            ws.cell(row=rr, column=1).border = BORDER
            sc = ws.cell(row=rr, column=2, value=i + 1)
            sc.font = f(size=9); sc.border = BORDER
            sc.alignment = Alignment(horizontal="center")
            ten, sl, gia = EX.get((d, i), (None, None, None)) if m == 1 else (None, None, None)
            for col, val, fmt in ((3, ten, None), (4, sl, '0'), (5, gia, VND)):
                cc = ws.cell(row=rr, column=col)
                if val is not None:
                    cc.value = val
                cc.font = f(blue=True, size=9)
                cc.fill = INPUT_FILL
                cc.border = BORDER
                if fmt:
                    cc.number_format = fmt
            fc = ws.cell(row=rr, column=6, value=f"=D{rr}*E{rr}")
            fc.font = f(size=9); fc.border = BORDER; fc.number_format = VND
        r = first + N_ROW
        ws.merge_cells(f"A{r}:E{r}")
        tc = ws.cell(row=r, column=1, value=f"TỔNG NGÀY {d}/{m}")
        tc.font = f(bold=True, size=9)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = DAY_FILL
            ws.cell(row=r, column=col).border = BORDER
        ec = ws.cell(row=r, column=6, value=f"=SUM(F{first}:F{r - 1})")
        ec.font = f(bold=True, size=9); ec.number_format = VND
        week_day_rows.append(r)
        r += 1
        if d in (7, 14, 21, 28) or d == ndays:
            k = len(week_rows) + 1
            ws.merge_cells(f"A{r}:E{r}")
            tc = ws.cell(row=r, column=1, value=f"🔶 TỔNG TUẦN {k} (ngày {week_first_day}–{d}/{m})")
            tc.font = f(bold=True, size=10, color="843C0C")
            tc.alignment = Alignment(horizontal="right", vertical="center")
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = WEEK_FILL
                ws.cell(row=r, column=col).border = Border(top=med, bottom=med, left=thin, right=thin)
            ec = ws.cell(row=r, column=6, value="=" + "+".join(f"F{x}" for x in week_day_rows))
            ec.font = f(bold=True, size=10, color="843C0C"); ec.number_format = VND
            ws.row_dimensions[r].height = 18
            week_rows.append(r)
            week_day_rows = []
            week_first_day = d + 1
            r += 1
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    tc = ws.cell(row=r, column=1, value=f"💰 TỔNG MUA HÀNG THÁNG {m}")
    tc.font = f(bold=True, size=12, white=True)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = GRAND_FILL
    ec = ws.cell(row=r, column=6, value="=" + "+".join(f"F{x}" for x in week_rows))
    ec.font = f(bold=True, size=12, white=True); ec.number_format = VND
    ws.row_dimensions[r].height = 24
    mh_refs[m] = (week_rows, r)

# ============ 12 TAB "Tháng m" (tong quan) ============
WEEK_LABEL = ["ngày 1–7", "ngày 8–14", "ngày 15–21", "ngày 22–28", "ngày 29–hết tháng"]
for idx in range(12):
    m = idx + 1
    week_rows, mh_total_row = mh_refs[m]
    mh = f"'MH Tháng {m}'"
    ws = wb.create_sheet(f"Tháng {m}")
    ws.sheet_properties.tabColor = TAB_COLORS[idx]
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 6, "B": 34, "C": 13, "D": 16, "E": 18}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"TỔNG QUAN THU CHI THÁNG {m}/2026"
    c.font = f(bold=True, size=15, white=True)
    c.fill = PatternFill("solid", fgColor=TAB_COLORS[idx])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    def section(row, text, color):
        ws.merge_cells(f"A{row}:E{row}")
        cc = ws.cell(row=row, column=1, value=text)
        cc.font = f(bold=True, size=11, white=True)
        cc.fill = PatternFill("solid", fgColor=color)
        cc.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 20

    def headers(row, labels):
        for j, lab in enumerate(labels, start=1):
            cc = ws.cell(row=row, column=j, value=lab)
            cc.font = f(bold=True, white=True, size=9)
            cc.fill = PatternFill("solid", fgColor=NAVY)
            cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cc.border = BORDER

    def stt(r, v):
        cc = ws.cell(row=r, column=1, value=v)
        cc.font = f(); cc.border = BORDER
        cc.alignment = Alignment(horizontal="center")

    def input_cell(r, col, value=None, numfmt=VND):
        cc = ws.cell(row=r, column=col)
        if value is not None:
            cc.value = value
        cc.font = f(blue=True)
        cc.fill = INPUT_FILL
        cc.border = BORDER
        if numfmt:
            cc.number_format = numfmt

    def total_row(r, label, formula, note):
        ws.merge_cells(f"A{r}:C{r}")
        tc = ws.cell(row=r, column=1, value=label)
        tc.font = f(bold=True); tc.alignment = Alignment(horizontal="right", vertical="center")
        nc = ws.cell(row=r, column=4, value=note)
        nc.font = Font(name=ARIAL, italic=True, size=8, color="808080")
        nc.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = TOTAL_FILL
            ws.cell(row=r, column=col).border = Border(top=med, bottom=med, left=thin, right=thin)
        ec = ws.cell(row=r, column=5, value=formula)
        ec.font = f(bold=True, size=11); ec.number_format = VND
        ws.row_dimensions[r].height = 20

    # ---- I. TIEN LUONG (15 dong) ----
    section(3, "I. TIỀN LƯƠNG NHÂN VIÊN", BLUE_SEC)
    headers(4, ["STT", "Khoản mục / Tên nhân viên", "Số lượng NV", "Lương/người (VNĐ)", "Thành tiền (VNĐ)"])
    for i in range(N_S):
        r = SAL_FIRST + i
        stt(r, i + 1)
        if i < len(STAFF):
            bc = ws.cell(row=r, column=2, value=STAFF[i])
            bc.font = f(); bc.border = BORDER
            if i % 2 == 1:
                bc.fill = BAND_FILL
                ws.cell(row=r, column=1).fill = BAND_FILL
        else:
            input_cell(r, 2, None, numfmt=None)
        sl, luong = EX_SALARY[i] if (m == 1 and i < len(EX_SALARY)) else (None, None)
        input_cell(r, 3, sl, numfmt='0')
        input_cell(r, 4, luong)
        cc = ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
        cc.font = f(); cc.border = BORDER; cc.number_format = VND
    total_row(ROW_TL, "TỔNG TIỀN LƯƠNG (1)",
              f"=SUM(E{SAL_FIRST}:E{SAL_LAST})", f"=SUM(E{SAL_FIRST}:E{SAL_LAST})")

    # ---- II. MUA HANG (tu keo theo tuan tu tab MH) ----
    section(SEC2, f"II. CHI PHÍ MUA HÀNG — tự kéo từ tab 'MH Tháng {m}' (nhập chi tiết theo ngày ở tab đó)", GREEN_SEC)
    headers(HDR2, ["Tuần", "Khoảng ngày", "", "", "Thành tiền (VNĐ)"])
    for k in range(5):
        r = WK_FIRST + k
        cc = ws.cell(row=r, column=1, value=f"Tuần {k+1}")
        cc.font = f(bold=True, size=9); cc.border = BORDER
        cc.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"B{r}:D{r}")
        bc = ws.cell(row=r, column=2, value=WEEK_LABEL[k] if (k < len(week_rows)) else "(tháng này không có tuần 5)")
        bc.font = f(size=9); bc.border = BORDER
        for col in (3, 4):
            ws.cell(row=r, column=col).border = BORDER
        ec = ws.cell(row=r, column=5)
        if k < len(week_rows):
            ec.value = f"={mh}!F{week_rows[k]}"
        ec.font = f(size=9); ec.border = BORDER; ec.number_format = VND
        if k % 2 == 1:
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = BAND_FILL
    total_row(ROW_MH, "TỔNG CHI PHÍ MUA HÀNG (2)",
              f"=SUM(E{WK_FIRST}:E{WK_FIRST+4})", f"=SUM(E{WK_FIRST}:E{WK_FIRST+4})")

    # ---- III. VAN HANH & PHAT SINH ----
    section(SEC3, "III. CHI PHÍ VẬN HÀNH & CHI PHÍ PHÁT SINH", ORANGE_SEC)
    headers(HDR3, ["STT", "Khoản mục", "Ghi chú", "", "Thành tiền (VNĐ)"])
    for i, label in enumerate(OPS):
        r = ROW_O1 + i
        stt(r, i + 1)
        bc = ws.cell(row=r, column=2, value=label)
        bc.font = f(); bc.border = BORDER
        if i % 2 == 1:
            bc.fill = BAND_FILL
            ws.cell(row=r, column=1).fill = BAND_FILL
        ws.merge_cells(f"C{r}:D{r}")
        input_cell(r, 3, None, numfmt=None)
        ws.cell(row=r, column=4).border = BORDER
        input_cell(r, 5, EX_OPS[i] if m == 1 else None)
    total_row(ROW_T3, "TỔNG VẬN HÀNH & PHÁT SINH (3)",
              f"=SUM(E{ROW_O1}:E{ROW_O1+6})", f"=SUM(E{ROW_O1}:E{ROW_O1+6})")

    # ---- TONG CHI PHI THANG ----
    g = ROW_G
    ws.merge_cells(f"A{g}:C{g}")
    tc = ws.cell(row=g, column=1, value=f"💰 TỔNG CHI PHÍ THÁNG {m}")
    tc.font = f(bold=True, size=13, white=True)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    nc = ws.cell(row=g, column=4, value="= (1) + (2) + (3)")
    nc.font = Font(name=ARIAL, italic=True, size=9, color="FFD966")
    nc.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=g, column=col).fill = GRAND_FILL
    ec = ws.cell(row=g, column=5, value=f"=E{ROW_TL}+E{ROW_MH}+E{ROW_T3}")
    ec.font = f(bold=True, size=13, white=True)
    ec.number_format = VND
    ws.row_dimensions[g].height = 26

    # ---- DOANH THU & LOI NHUAN ----
    r = ROW_DT
    ws.merge_cells(f"A{r}:C{r}")
    tc = ws.cell(row=r, column=1, value=f"📈 DOANH THU THÁNG {m}")
    tc.font = f(bold=True, size=12, white=True)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    nc = ws.cell(row=r, column=4, value="(bạn nhập vào ô vàng →)")
    nc.font = Font(name=ARIAL, italic=True, size=9, color="E2EFDA")
    nc.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 5):
        ws.cell(row=r, column=col).fill = REV_FILL
    dc = ws.cell(row=r, column=5)
    if m == 1:
        dc.value = 250000000
    dc.font = Font(name=ARIAL, bold=True, size=12, color="0D47A1")
    dc.fill = INPUT_FILL
    dc.border = BORDER
    dc.number_format = VND
    ws.row_dimensions[r].height = 24

    r = ROW_LN
    ws.merge_cells(f"A{r}:C{r}")
    tc = ws.cell(row=r, column=1, value=f"✅ LỢI NHUẬN THÁNG {m}")
    tc.font = f(bold=True, size=13, white=True)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    nc = ws.cell(row=r, column=4, value="= Doanh thu − Tổng chi phí")
    nc.font = Font(name=ARIAL, italic=True, size=9, color="C6E0B4")
    nc.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=r, column=col).fill = GREEN_FILL
    lc = ws.cell(row=r, column=5, value=f"=E{ROW_DT}-E{ROW_G}")
    lc.font = f(bold=True, size=13, white=True)
    lc.number_format = VND
    ws.row_dimensions[r].height = 26

# ============ TONG HOP NAM ============
ws = wb.create_sheet("TỔNG HỢP NĂM")
ws.sheet_properties.tabColor = "FFB300"
ws.sheet_view.showGridLines = False
for col, w in {"A": 12, "B": 17, "C": 19, "D": 20, "E": 16, "F": 16, "G": 19, "H": 18, "I": 19}.items():
    ws.column_dimensions[col].width = w
ws.merge_cells("A1:I1")
c = ws["A1"]
c.value = "TỔNG HỢP THU CHI & LỢI NHUẬN NĂM 2026"
c.font = f(bold=True, size=16, white=True)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32
ws.merge_cells("A2:I2")
c2 = ws.cell(row=2, column=1, value="Tự động cộng từ các tab tháng — không cần nhập gì ở trang này")
c2.font = Font(name=ARIAL, italic=True, size=9, color="808080")
c2.alignment = Alignment(horizontal="center")

labels = ["Tháng", "Tiền lương\n(VNĐ)", "Chi phí mua hàng\n(VNĐ)", "Điện + nước + mạng\n(VNĐ)",
          "Quảng cáo\n(VNĐ)", "Phát sinh\n(VNĐ)", "TỔNG CHI PHÍ\n(VNĐ)", "Doanh thu\n(VNĐ)", "LỢI NHUẬN\n(VNĐ)"]
for j, lab in enumerate(labels, start=1):
    cc = ws.cell(row=4, column=j, value=lab)
    cc.font = f(bold=True, white=True, size=9)
    cc.fill = PatternFill("solid", fgColor=NAVY)
    cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cc.border = BORDER
ws.row_dimensions[4].height = 30

for m in range(1, 13):
    r = 4 + m
    t = f"'Tháng {m}'"
    vals = [f"Tháng {m}",
            f"={t}!E{ROW_TL}",
            f"={t}!E{ROW_MH}",
            f"={t}!E{ROW_O1}+{t}!E{ROW_O1+1}+{t}!E{ROW_O1+2}",
            f"={t}!E{ROW_O1+3}",
            f"=SUM({t}!E{ROW_O1+4}:E{ROW_O1+6})",
            f"=SUM(B{r}:F{r})",
            f"={t}!E{ROW_DT}",
            f"=H{r}-G{r}"]
    for j, v in enumerate(vals, start=1):
        cc = ws.cell(row=r, column=j, value=v)
        cc.border = BORDER
        if j == 1:
            cc.alignment = Alignment(horizontal="center")
            cc.fill = PatternFill("solid", fgColor=TAB_COLORS[m-1])
            cc.font = f(bold=True, white=True)
        else:
            cc.font = f(bold=(j in (7, 9)))
            cc.number_format = VND
        if j == 9:
            cc.fill = PatternFill("solid", fgColor="E2EFDA")
        elif m % 2 == 0 and j > 1:
            cc.fill = BAND_FILL

r = 17
ws.row_dimensions[r].height = 26
cc = ws.cell(row=r, column=1, value="💰 CẢ NĂM")
cc.font = f(bold=True, size=12, white=True)
cc.alignment = Alignment(horizontal="center", vertical="center")
for j in range(1, 10):
    col = get_column_letter(j)
    cell = ws.cell(row=r, column=j)
    if j > 1:
        cell.value = f"=SUM({col}5:{col}16)"
        cell.number_format = VND
        cell.font = f(bold=True, size=12, white=True)
    cell.fill = GREEN_FILL if j == 9 else GRAND_FILL
    cell.border = Border(top=med, bottom=med, left=thin, right=thin)

# ============ SAP XEP THU TU TAB ============
order = ["HƯỚNG DẪN"]
for m in range(1, 13):
    order += [f"Tháng {m}", f"MH Tháng {m}"]
order += ["TỔNG HỢP NĂM"]
wb._sheets = [wb[name] for name in order]

out = "/tmp/claude-0/-home-user-lien-ket-tiktok/26731859-2296-588b-8705-3fe299e2bd5a/scratchpad/Quan-ly-thu-chi-loi-nhuan-2026-Bao-Qua-Tang.xlsx"
wb.save(out)

# ---- kiem tra ----
wb2 = load_workbook(out)
t1 = wb2["Tháng 1"]
mh1 = wb2["MH Tháng 1"]
th = wb2["TỔNG HỢP NĂM"]
w1, mrow1 = mh_refs[1]
w2, _ = mh_refs[2]
checks = [
    ("MH1 F4", mh1["F4"].value, "=D4*E4"),
    ("MH1 F12 tong ngay 1", mh1["F12"].value, "=SUM(F4:F11)"),
    ("MH1 tuan 1", mh1[f"F{w1[0]}"].value, "=F12+F21+F30+F39+F48+F57+F66"),
    ("MH1 tong thang", mh1[f"F{mrow1}"].value, "=" + "+".join(f"F{x}" for x in w1)),
    ("T1 E20 tong luong", t1["E20"].value, "=SUM(E5:E19)"),
    ("T1 E24 tuan 1", t1["E24"].value, f"='MH Tháng 1'!F{w1[0]}"),
    ("T1 E28 tuan 5", t1["E28"].value, f"='MH Tháng 1'!F{w1[4]}"),
    ("T1 E29 tong mua hang", t1["E29"].value, "=SUM(E24:E28)"),
    ("T1 E40 tong van hanh", t1["E40"].value, "=SUM(E33:E39)"),
    ("T1 E42 tong chi phi", t1["E42"].value, "=E20+E29+E40"),
    ("T1 E45 loi nhuan", t1["E45"].value, "=E44-E42"),
    ("T2 E28 tuan 5 (trong)", wb2["Tháng 2"]["E28"].value, None),
    ("T2 so tuan", len(w2), 4),
    ("TH C5 mua hang", th["C5"].value, "='Tháng 1'!E29"),
    ("TH G5 tong chi phi", th["G5"].value, "=SUM(B5:F5)"),
    ("TH I5 loi nhuan", th["I5"].value, "=H5-G5"),
    ("TH I17 loi nhuan nam", th["I17"].value, "=SUM(I5:I16)"),
]
ok = True
for name, got, want in checks:
    status = "OK " if got == want else "FAIL"
    if got != want: ok = False
    print(f"{status} {name}: {got}" + ("" if got == want else f"  (mong doi: {want})"))
print("so tab:", len(wb2.sheetnames))
print("thu tu:", wb2.sheetnames[:5], "...", wb2.sheetnames[-3:])
print("ALL OK" if ok else "CO LOI!")
print("size:", os.path.getsize(out), "bytes")
