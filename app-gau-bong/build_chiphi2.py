#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""File quan ly chi phi 2026 — 12 tab thang + tab tong hop nam.
Phan I: 15 dong luong. Phan II: 30 dong mua hang."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
NAVY = "1F3864"
BLUE_SEC = "2E75B6"
GREEN_SEC = "548235"
ORANGE_SEC = "C55A11"
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
GRAND_FILL = PatternFill("solid", fgColor="1F3864")
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
med = Side(style="medium", color="1F3864")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
VND = '#,##0'

TAB_COLORS = ["E53935", "F4511E", "FB8C00", "FDD835", "7CB342", "43A047",
              "00897B", "039BE5", "3949AB", "5E35B1", "8E24AA", "D81B60"]

N_S = 15   # so dong tien luong
N_P = 30   # so dong mua hang

# layout tinh san (giong nhau moi tab thang)
SAL_FIRST = 5
SAL_LAST = 4 + N_S                 # 19
ROW_TL = SAL_LAST + 1              # 20  tong luong
SEC2 = ROW_TL + 2                  # 22
HDR2 = SEC2 + 1                    # 23
P_FIRST = HDR2 + 1                 # 24
P_LAST = HDR2 + N_P                # 53
ROW_MH = P_LAST + 1                # 54  tong mua hang
SEC3 = ROW_MH + 2                  # 56
HDR3 = SEC3 + 1                    # 57
ROW_O1 = HDR3 + 1                  # 58  dong Tien dien
ROW_T3 = ROW_O1 + 7                # 65  tong van hanh
ROW_G = ROW_T3 + 2                 # 67  tong thang
ROW_DT = ROW_G + 2                 # 69  doanh thu thang (nhap tay)
ROW_LN = ROW_DT + 1                # 70  loi nhuan = doanh thu - tong chi phi
GREEN_FILL = PatternFill("solid", fgColor="375623")   # nen loi nhuan
REV_FILL = PatternFill("solid", fgColor="548235")     # nen doanh thu

def f(bold=False, size=10, color="000000", white=False, blue=False):
    c = "FFFFFF" if white else ("0D47A1" if blue else color)
    return Font(name=ARIAL, bold=bold, size=size, color=c)

wb = Workbook()

# ============ HUONG DAN ============
ws = wb.active
ws.title = "HƯỚNG DẪN"
ws.sheet_properties.tabColor = "757575"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 105
guide = [
    ("HƯỚNG DẪN SỬ DỤNG — FILE QUẢN LÝ CHI PHÍ 2026", 1),
    ("", 0),
    ("• File có 12 tab 'Tháng 1' → 'Tháng 12' (mỗi tháng 1 trang riêng, mỗi tab 1 màu) và tab cuối 'TỔNG HỢP NĂM'.", 0),
    ("• Mỗi tháng gồm 3 phần:  I. Tiền lương nhân viên (15 dòng)  ·  II. Chi phí mua hàng (30 dòng)  ·  III. Chi phí vận hành & phát sinh.", 0),
    ("• Bạn CHỈ nhập vào các ô NỀN VÀNG NHẠT (chữ xanh). Các ô khác có công thức tự tính — không sửa tay.", 0),
    ("• Phần I: 4 dòng đầu đã ghi sẵn vị trí (marketing, sale, đóng hàng, sản xuất); 11 dòng sau để trống — bạn tự điền thêm tên vị trí/nhân viên.", 0),
    ("", 0),
    ("CÔNG THỨC TRONG FILE (tự động, không cần đụng vào):", 2),
    ("   - Thành tiền  =  Số lượng  ×  Đơn giá   (ví dụ ô E5 = C5*D5)", 0),
    ("   - TỔNG mỗi phần  =  SUM(các dòng trong phần đó)", 0),
    ("   - TỔNG CHI PHÍ THÁNG (ô cuối mỗi tab)  =  Tổng (1) + Tổng (2) + Tổng (3)", 0),
    ("   - DOANH THU THÁNG: bạn nhập vào ô vàng cuối mỗi tab tháng.", 0),
    ("   - LỢI NHUẬN THÁNG  =  Doanh thu  −  Tổng chi phí tháng   (tự tính, dòng xanh lá cuối tab)", 0),
    ("   - Tab TỔNG HỢP NĂM tự kéo số từ 12 tab tháng: chi phí, doanh thu, LỢI NHUẬN từng tháng và CẢ NĂM.", 0),
    ("", 0),
    ("• Tháng 1 đã điền SẴN SỐ VÍ DỤ để bạn xem cách nhập — hãy thay bằng số thật của bạn.", 0),
    ("• Đơn vị tiền: VNĐ.", 0),
]
for i, (text, style) in enumerate(guide, start=1):
    c = ws.cell(row=i, column=2, value=text)
    c.font = f(bold=(style > 0), size=14 if style == 1 else 11)
    if style == 1:
        c.font = Font(name=ARIAL, bold=True, size=14, color=NAVY)

MONTHS = [f"Tháng {m}" for m in range(1, 13)]
EX_SALARY = [(2, 9000000), (3, 8000000), (2, 7000000), (4, 8500000)]
EX_PURCHASE = [("Hộp quà tặng (ví dụ)", 200, 25000), ("Túi giấy in logo (ví dụ)", 500, 8000)]
EX_OPS = [3500000, 500000, 400000, 15000000, 1200000, None, None]
STAFF = ["Nhân viên marketing", "Nhân viên sale", "Nhân viên đóng hàng", "Nhân viên sản xuất"]
OPS = ["Tiền điện", "Tiền nước", "Tiền mạng (internet)", "Tiền chạy quảng cáo",
       "Chi phí phát sinh 1", "Chi phí phát sinh 2", "Chi phí phát sinh 3"]

for idx, name in enumerate(MONTHS):
    m = idx + 1
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLORS[idx]
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 6, "B": 34, "C": 13, "D": 16, "E": 18}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"BẢNG CHI PHÍ THÁNG {m}/2026"
    c.font = f(bold=True, size=15, white=True)
    c.fill = PatternFill("solid", fgColor=TAB_COLORS[idx])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    def section(row, text, color):
        ws.merge_cells(f"A{row}:E{row}")
        cc = ws.cell(row=row, column=1, value=text)
        cc.font = f(bold=True, size=11, white=True)
        cc.fill = PatternFill("solid", fgColor=color)
        cc.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 20

    def headers(row, labels):
        for j, lab in enumerate(labels, start=1):
            cc = ws.cell(row=row, column=j, value=lab)
            cc.font = f(bold=True, white=True, size=9)
            cc.fill = PatternFill("solid", fgColor=NAVY)
            cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cc.border = BORDER

    def stt(r, v):
        cc = ws.cell(row=r, column=1, value=v)
        cc.font = f(); cc.border = BORDER
        cc.alignment = Alignment(horizontal="center")

    def input_cell(r, col, value=None, numfmt=VND):
        cc = ws.cell(row=r, column=col)
        if value is not None:
            cc.value = value
        cc.font = f(blue=True)
        cc.fill = INPUT_FILL
        cc.border = BORDER
        if numfmt:
            cc.number_format = numfmt

    def formula_cell(r, col, formula, bold=False):
        cc = ws.cell(row=r, column=col, value=formula)
        cc.font = f(bold=bold); cc.border = BORDER
        cc.number_format = VND

    def total_row(r, label, formula, note):
        ws.merge_cells(f"A{r}:C{r}")
        tc = ws.cell(row=r, column=1, value=label)
        tc.font = f(bold=True); tc.alignment = Alignment(horizontal="right", vertical="center")
        nc = ws.cell(row=r, column=4, value=note)
        nc.font = Font(name=ARIAL, italic=True, size=8, color="808080")
        nc.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = TOTAL_FILL
            ws.cell(row=r, column=col).border = Border(top=med, bottom=med, left=thin, right=thin)
        ec = ws.cell(row=r, column=5, value=formula)
        ec.font = f(bold=True, size=11); ec.number_format = VND
        ws.row_dimensions[r].height = 20

    # ---- I. TIEN LUONG (15 dong) ----
    section(3, "I. TIỀN LƯƠNG NHÂN VIÊN", BLUE_SEC)
    headers(4, ["STT", "Khoản mục / Tên nhân viên", "Số lượng NV", "Lương/người (VNĐ)", "Thành tiền (VNĐ)"])
    for i in range(N_S):
        r = SAL_FIRST + i
        stt(r, i + 1)
        if i < len(STAFF):
            bc = ws.cell(row=r, column=2, value=STAFF[i])
            bc.font = f(); bc.border = BORDER
            if i % 2 == 1:
                bc.fill = BAND_FILL
                ws.cell(row=r, column=1).fill = BAND_FILL
        else:
            input_cell(r, 2, None, numfmt=None)   # o vang: tu dien vi tri/ten
        sl, luong = EX_SALARY[i] if (m == 1 and i < len(EX_SALARY)) else (None, None)
        input_cell(r, 3, sl, numfmt='0')
        input_cell(r, 4, luong)
        formula_cell(r, 5, f"=C{r}*D{r}")
    total_row(ROW_TL, "TỔNG TIỀN LƯƠNG (1)",
              f"=SUM(E{SAL_FIRST}:E{SAL_LAST})", f"=SUM(E{SAL_FIRST}:E{SAL_LAST})")

    # ---- II. MUA HANG (30 dong) ----
    section(SEC2, "II. CHI PHÍ MUA HÀNG", GREEN_SEC)
    headers(HDR2, ["STT", "Tên sản phẩm", "Số lượng", "Đơn giá (VNĐ)", "Thành tiền (VNĐ)"])
    for i in range(N_P):
        r = P_FIRST + i
        stt(r, i + 1)
        ten, sl, gia = EX_PURCHASE[i] if (m == 1 and i < len(EX_PURCHASE)) else (None, None, None)
        input_cell(r, 2, ten, numfmt=None)
        input_cell(r, 3, sl, numfmt='0')
        input_cell(r, 4, gia)
        formula_cell(r, 5, f"=C{r}*D{r}")
    total_row(ROW_MH, "TỔNG CHI PHÍ MUA HÀNG (2)",
              f"=SUM(E{P_FIRST}:E{P_LAST})", f"=SUM(E{P_FIRST}:E{P_LAST})")

    # ---- III. VAN HANH & PHAT SINH ----
    section(SEC3, "III. CHI PHÍ VẬN HÀNH & CHI PHÍ PHÁT SINH", ORANGE_SEC)
    headers(HDR3, ["STT", "Khoản mục", "Ghi chú", "", "Thành tiền (VNĐ)"])
    for i, label in enumerate(OPS):
        r = ROW_O1 + i
        stt(r, i + 1)
        bc = ws.cell(row=r, column=2, value=label)
        bc.font = f(); bc.border = BORDER
        if i % 2 == 1:
            bc.fill = BAND_FILL
            ws.cell(row=r, column=1).fill = BAND_FILL
        ws.merge_cells(f"C{r}:D{r}")
        input_cell(r, 3, None, numfmt=None)
        ws.cell(row=r, column=4).border = BORDER
        input_cell(r, 5, EX_OPS[i] if m == 1 else None)
    total_row(ROW_T3, "TỔNG VẬN HÀNH & PHÁT SINH (3)",
              f"=SUM(E{ROW_O1}:E{ROW_O1+6})", f"=SUM(E{ROW_O1}:E{ROW_O1+6})")

    # ---- TONG THANG (o cuoi tab) ----
    g = ROW_G
    ws.merge_cells(f"A{g}:C{g}")
    tc = ws.cell(row=g, column=1, value=f"💰 TỔNG CHI PHÍ THÁNG {m}")
    tc.font = f(bold=True, size=13, white=True)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    nc = ws.cell(row=g, column=4, value="= (1) + (2) + (3)")
    nc.font = Font(name=ARIAL, italic=True, size=9, color="FFD966")
    nc.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=g, column=col).fill = GRAND_FILL
    ec = ws.cell(row=g, column=5, value=f"=E{ROW_TL}+E{ROW_MH}+E{ROW_T3}")
    ec.font = f(bold=True, size=13, white=True)
    ec.number_format = VND
    ws.row_dimensions[g].height = 26

    # ---- DOANH THU & LOI NHUAN THANG ----
    r = ROW_DT
    ws.merge_cells(f"A{r}:C{r}")
    tc = ws.cell(row=r, column=1, value=f"📈 DOANH THU THÁNG {m}")
    tc.font = f(bold=True, size=12, white=True)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    nc = ws.cell(row=r, column=4, value="(bạn nhập vào ô vàng →)")
    nc.font = Font(name=ARIAL, italic=True, size=9, color="E2EFDA")
    nc.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 5):
        ws.cell(row=r, column=col).fill = REV_FILL
    dc = ws.cell(row=r, column=5)
    if m == 1:
        dc.value = 250000000   # vi du
    dc.font = Font(name=ARIAL, bold=True, size=12, color="0D47A1")
    dc.fill = INPUT_FILL
    dc.border = BORDER
    dc.number_format = VND
    ws.row_dimensions[r].height = 24

    r = ROW_LN
    ws.merge_cells(f"A{r}:C{r}")
    tc = ws.cell(row=r, column=1, value=f"✅ LỢI NHUẬN THÁNG {m}")
    tc.font = f(bold=True, size=13, white=True)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    nc = ws.cell(row=r, column=4, value="= Doanh thu − Tổng chi phí")
    nc.font = Font(name=ARIAL, italic=True, size=9, color="C6E0B4")
    nc.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=r, column=col).fill = GREEN_FILL
    lc = ws.cell(row=r, column=5, value=f"=E{ROW_DT}-E{ROW_G}")
    lc.font = f(bold=True, size=13, white=True)
    lc.number_format = VND
    ws.row_dimensions[r].height = 26

# ============ TONG HOP NAM (tab cuoi) ============
ws = wb.create_sheet("TỔNG HỢP NĂM")
ws.sheet_properties.tabColor = "FFB300"
ws.sheet_view.showGridLines = False
for col, w in {"A": 12, "B": 17, "C": 19, "D": 20, "E": 16, "F": 16, "G": 19, "H": 18, "I": 19}.items():
    ws.column_dimensions[col].width = w
ws.merge_cells("A1:I1")
c = ws["A1"]
c.value = "TỔNG HỢP CHI PHÍ NĂM 2026"
c.font = f(bold=True, size=16, white=True)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32
ws.merge_cells("A2:I2")
c2 = ws.cell(row=2, column=1, value="Tự động cộng từ 12 tab tháng — không cần nhập gì ở trang này")
c2.font = Font(name=ARIAL, italic=True, size=9, color="808080")
c2.alignment = Alignment(horizontal="center")

labels = ["Tháng", "Tiền lương\n(VNĐ)", "Chi phí mua hàng\n(VNĐ)", "Điện + nước + mạng\n(VNĐ)",
          "Quảng cáo\n(VNĐ)", "Phát sinh\n(VNĐ)", "TỔNG CHI PHÍ\n(VNĐ)", "Doanh thu\n(VNĐ)", "LỢI NHUẬN\n(VNĐ)"]
for j, lab in enumerate(labels, start=1):
    cc = ws.cell(row=4, column=j, value=lab)
    cc.font = f(bold=True, white=True, size=9)
    cc.fill = PatternFill("solid", fgColor=NAVY)
    cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cc.border = BORDER
ws.row_dimensions[4].height = 30

for m in range(1, 13):
    r = 4 + m
    t = f"'Tháng {m}'"
    vals = [f"Tháng {m}",
            f"={t}!E{ROW_TL}",
            f"={t}!E{ROW_MH}",
            f"={t}!E{ROW_O1}+{t}!E{ROW_O1+1}+{t}!E{ROW_O1+2}",
            f"={t}!E{ROW_O1+3}",
            f"=SUM({t}!E{ROW_O1+4}:E{ROW_O1+6})",
            f"=SUM(B{r}:F{r})",
            f"={t}!E{ROW_DT}",
            f"=H{r}-G{r}"]
    for j, v in enumerate(vals, start=1):
        cc = ws.cell(row=r, column=j, value=v)
        cc.border = BORDER
        if j == 1:
            cc.alignment = Alignment(horizontal="center")
            cc.fill = PatternFill("solid", fgColor=TAB_COLORS[m-1])
            cc.font = f(bold=True, white=True)
        else:
            cc.font = f(bold=(j in (7, 9)))
            cc.number_format = VND
        if j == 9:
            cc.fill = PatternFill("solid", fgColor="E2EFDA")   # cot loi nhuan xanh nhat
        elif m % 2 == 0 and j > 1:
            cc.fill = BAND_FILL

r = 17
ws.row_dimensions[r].height = 26
cc = ws.cell(row=r, column=1, value="💰 CẢ NĂM")
cc.font = f(bold=True, size=12, white=True)
cc.alignment = Alignment(horizontal="center", vertical="center")
for j in range(1, 10):
    col = get_column_letter(j)
    cell = ws.cell(row=r, column=j)
    if j > 1:
        cell.value = f"=SUM({col}5:{col}16)"
        cell.number_format = VND
        cell.font = f(bold=True, size=12, white=True)
    cell.fill = GREEN_FILL if j == 9 else GRAND_FILL
    cell.border = Border(top=med, bottom=med, left=thin, right=thin)

out = "/tmp/claude-0/-home-user-lien-ket-tiktok/26731859-2296-588b-8705-3fe299e2bd5a/scratchpad/Quan-ly-chi-phi-2026-Bao-Qua-Tang.xlsx"
wb.save(out)

# ---- doc nguoc kiem tra cong thuc ----
from openpyxl import load_workbook
wb2 = load_workbook(out)
t1 = wb2["Tháng 1"]
th = wb2["TỔNG HỢP NĂM"]
checks = [
    ("E5 (luong dong 1)", t1["E5"].value, "=C5*D5"),
    ("E19 (luong dong 15)", t1["E19"].value, "=C19*D19"),
    ("E20 (tong luong)", t1["E20"].value, "=SUM(E5:E19)"),
    ("E24 (mua hang dong 1)", t1["E24"].value, "=C24*D24"),
    ("E53 (mua hang dong 30)", t1["E53"].value, "=C53*D53"),
    ("E54 (tong mua hang)", t1["E54"].value, "=SUM(E24:E53)"),
    ("E65 (tong van hanh)", t1["E65"].value, "=SUM(E58:E64)"),
    ("E67 (tong thang)", t1["E67"].value, "=E20+E54+E65"),
    ("TH B5", th["B5"].value, "='Tháng 1'!E20"),
    ("TH C5", th["C5"].value, "='Tháng 1'!E54"),
    ("TH D5", th["D5"].value, "='Tháng 1'!E58+'Tháng 1'!E59+'Tháng 1'!E60"),
    ("TH G5", th["G5"].value, "=SUM(B5:F5)"),
    ("TH B17 (ca nam)", th["B17"].value, "=SUM(B5:B16)"),
    ("E70 (loi nhuan T1)", t1["E70"].value, "=E69-E67"),
    ("E69 (doanh thu vi du)", t1["E69"].value, 250000000),
    ("TH H5 (doanh thu)", th["H5"].value, "='Tháng 1'!E69"),
    ("TH I5 (loi nhuan)", th["I5"].value, "=H5-G5"),
    ("TH I17 (loi nhuan nam)", th["I17"].value, "=SUM(I5:I16)"),
]
ok = True
for name, got, want in checks:
    status = "OK " if got == want else "FAIL"
    if got != want: ok = False
    print(f"{status} {name}: {got}")
print("sheets:", wb2.sheetnames)
print("ALL OK" if ok else "CO LOI!")
import os
print("size:", os.path.getsize(out), "bytes")
