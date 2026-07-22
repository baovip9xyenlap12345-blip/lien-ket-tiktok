#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""File CHI PHI MUA HANG THEO NGAY 2026:
- 12 tab thang, moi ngay 8 dong nhap hang
- Tong theo NGAY, tong theo TUAN (T1:1-7, T2:8-14, T3:15-21, T4:22-28, T5:29-het), tong THANG
- Tab TONG HOP NAM: chia theo tuan tung thang + ca nam"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

ARIAL = "Arial"
NAVY = "1F3864"
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
DAY_FILL = PatternFill("solid", fgColor="DDEBF7")     # tong ngay
WEEK_FILL = PatternFill("solid", fgColor="FCE4D6")    # tong tuan
GRAND_FILL = PatternFill("solid", fgColor="1F3864")   # tong thang
BAND_FILL = PatternFill("solid", fgColor="EDEDED")    # o ngay xen ke
thin = Side(style="thin", color="BFBFBF")
med = Side(style="medium", color="1F3864")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
VND = '#,##0'
TAB_COLORS = ["E53935", "F4511E", "FB8C00", "FDD835", "7CB342", "43A047",
              "00897B", "039BE5", "3949AB", "5E35B1", "8E24AA", "D81B60"]
DAYS = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
N_ROW = 8  # so dong hang moi ngay

def f(bold=False, size=10, white=False, blue=False, color="000000"):
    c = "FFFFFF" if white else ("0D47A1" if blue else color)
    return Font(name=ARIAL, bold=bold, size=size, color=c)

wb = Workbook()

# ============ HUONG DAN ============
ws = wb.active
ws.title = "HƯỚNG DẪN"
ws.sheet_properties.tabColor = "757575"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 110
guide = [
    ("HƯỚNG DẪN — FILE CHI PHÍ MUA HÀNG THEO NGÀY 2026", 1),
    ("", 0),
    ("• File có 12 tab 'Tháng 1' → 'Tháng 12' và tab cuối 'TỔNG HỢP NĂM'.", 0),
    ("• Mỗi NGÀY có sẵn 8 dòng để nhập hàng mua trong ngày (mua nhiều món cùng lúc).", 0),
    ("• Bạn CHỈ nhập vào ô NỀN VÀNG: Tên sản phẩm, Số lượng, Đơn giá. Thành tiền tự nhân = Số lượng × Đơn giá.", 0),
    ("", 0),
    ("CÁC DÒNG TỰ TÍNH (không sửa tay):", 2),
    ("   - TỔNG NGÀY (nền xanh nhạt)  =  SUM(8 dòng của ngày đó)", 0),
    ("   - TỔNG TUẦN (nền cam nhạt):  Tuần 1 = ngày 1–7 · Tuần 2 = ngày 8–14 · Tuần 3 = ngày 15–21 · Tuần 4 = ngày 22–28 · Tuần 5 = ngày 29–cuối tháng", 0),
    ("   - 💰 TỔNG CHI PHÍ MUA HÀNG THÁNG (nền xanh đậm, cuối tab)  =  cộng 5 tuần", 0),
    ("   - Tab TỔNG HỢP NĂM: chi phí từng tuần của từng tháng + tổng tháng + CẢ NĂM.", 0),
    ("", 0),
    ("• Ngày 1 và ngày 2 của Tháng 1 có SỐ VÍ DỤ (chữ xanh) — thay bằng số thật của bạn.", 0),
    ("• Ngày nào không mua hàng thì để trống, tổng tự = 0. Đơn vị tiền: VNĐ.", 0),
]
for i, (text, style) in enumerate(guide, start=1):
    c = ws.cell(row=i, column=2, value=text)
    c.font = Font(name=ARIAL, bold=True, size=14, color=NAVY) if style == 1 else f(bold=(style > 0), size=11)

EX = {  # (ngay, dong) -> (ten, sl, gia)  — vi du Thang 1
    (1, 0): ("Hộp quà tặng (ví dụ)", 200, 25000),
    (1, 1): ("Túi giấy in logo (ví dụ)", 500, 8000),
    (1, 2): ("Ruy băng (ví dụ)", 20, 35000),
    (2, 0): ("Bình giữ nhiệt in logo (ví dụ)", 100, 95000),
}

month_refs = {}  # m -> (week_rows[5], month_row)

for idx in range(12):
    m = idx + 1
    ndays = DAYS[m]
    ws = wb.create_sheet(f"Tháng {m}")
    ws.sheet_properties.tabColor = TAB_COLORS[idx]
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 11, "B": 5, "C": 34, "D": 11, "E": 14, "F": 17}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"CHI PHÍ MUA HÀNG THEO NGÀY — THÁNG {m}/2026"
    c.font = f(bold=True, size=14, white=True)
    c.fill = PatternFill("solid", fgColor=TAB_COLORS[idx])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for j, lab in enumerate(["Ngày", "STT", "Tên sản phẩm", "Số lượng", "Đơn giá (VNĐ)", "Thành tiền (VNĐ)"], start=1):
        cc = ws.cell(row=3, column=j, value=lab)
        cc.font = f(bold=True, white=True, size=9)
        cc.fill = PatternFill("solid", fgColor=NAVY)
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cc.border = BORDER
    ws.row_dimensions[3].height = 20

    r = 4
    day_rows = []          # dong TONG NGAY
    week_rows = []         # dong TONG TUAN
    week_first_day = 1
    week_day_rows = []
    for d in range(1, ndays + 1):
        first = r
        # 8 dong nhap hang
        ws.merge_cells(f"A{first}:A{first + N_ROW - 1}")
        ac = ws.cell(row=first, column=1, value=f"Ngày {d}/{m}")
        ac.font = f(bold=True, size=9)
        ac.alignment = Alignment(horizontal="center", vertical="center")
        ac.fill = BAND_FILL if d % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for i in range(N_ROW):
            rr = first + i
            ws.cell(row=rr, column=1).border = BORDER
            sc = ws.cell(row=rr, column=2, value=i + 1)
            sc.font = f(size=9); sc.border = BORDER
            sc.alignment = Alignment(horizontal="center")
            ten, sl, gia = EX.get((d, i), (None, None, None)) if m == 1 else (None, None, None)
            for col, val, fmt in ((3, ten, None), (4, sl, '0'), (5, gia, VND)):
                cc = ws.cell(row=rr, column=col)
                if val is not None:
                    cc.value = val
                cc.font = f(blue=True, size=9)
                cc.fill = INPUT_FILL
                cc.border = BORDER
                if fmt:
                    cc.number_format = fmt
            fc = ws.cell(row=rr, column=6, value=f"=D{rr}*E{rr}")
            fc.font = f(size=9); fc.border = BORDER; fc.number_format = VND
        r = first + N_ROW
        # dong tong ngay
        ws.merge_cells(f"A{r}:E{r}")
        tc = ws.cell(row=r, column=1, value=f"TỔNG NGÀY {d}/{m}")
        tc.font = f(bold=True, size=9)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = DAY_FILL
            ws.cell(row=r, column=col).border = BORDER
        ec = ws.cell(row=r, column=6, value=f"=SUM(F{first}:F{r - 1})")
        ec.font = f(bold=True, size=9); ec.number_format = VND
        day_rows.append(r)
        week_day_rows.append(r)
        r += 1
        # dong tong tuan
        if d in (7, 14, 21, 28) or d == ndays:
            k = len(week_rows) + 1
            ws.merge_cells(f"A{r}:E{r}")
            tc = ws.cell(row=r, column=1, value=f"🔶 TỔNG TUẦN {k} (ngày {week_first_day}–{d}/{m})")
            tc.font = f(bold=True, size=10, color="843C0C")
            tc.alignment = Alignment(horizontal="right", vertical="center")
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = WEEK_FILL
                ws.cell(row=r, column=col).border = Border(top=med, bottom=med, left=thin, right=thin)
            ec = ws.cell(row=r, column=6, value="=" + "+".join(f"F{x}" for x in week_day_rows))
            ec.font = f(bold=True, size=10, color="843C0C"); ec.number_format = VND
            ws.row_dimensions[r].height = 18
            week_rows.append(r)
            week_day_rows = []
            week_first_day = d + 1
            r += 1
    # tong thang
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    tc = ws.cell(row=r, column=1, value=f"💰 TỔNG CHI PHÍ MUA HÀNG THÁNG {m} (= 5 tuần cộng lại)")
    tc.font = f(bold=True, size=12, white=True)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = GRAND_FILL
    ec = ws.cell(row=r, column=6, value="=" + "+".join(f"F{x}" for x in week_rows))
    ec.font = f(bold=True, size=12, white=True); ec.number_format = VND
    ws.row_dimensions[r].height = 24
    month_refs[m] = (week_rows, r)

# ============ TONG HOP NAM ============
ws = wb.create_sheet("TỔNG HỢP NĂM")
ws.sheet_properties.tabColor = "FFB300"
ws.sheet_view.showGridLines = False
for col, w in {"A": 12, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 20}.items():
    ws.column_dimensions[col].width = w
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "TỔNG HỢP CHI PHÍ MUA HÀNG NĂM 2026 (THEO TUẦN)"
c.font = f(bold=True, size=14, white=True)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30
ws.merge_cells("A2:G2")
c2 = ws.cell(row=2, column=1, value="Tự động kéo số từ 12 tab tháng — không cần nhập gì ở trang này")
c2.font = Font(name=ARIAL, italic=True, size=9, color="808080")
c2.alignment = Alignment(horizontal="center")

labels = ["Tháng", "Tuần 1\n(ngày 1–7)", "Tuần 2\n(ngày 8–14)", "Tuần 3\n(ngày 15–21)",
          "Tuần 4\n(ngày 22–28)", "Tuần 5\n(ngày 29–hết)", "TỔNG THÁNG\n(VNĐ)"]
for j, lab in enumerate(labels, start=1):
    cc = ws.cell(row=4, column=j, value=lab)
    cc.font = f(bold=True, white=True, size=9)
    cc.fill = PatternFill("solid", fgColor=NAVY)
    cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cc.border = BORDER
ws.row_dimensions[4].height = 30

for m in range(1, 13):
    r = 4 + m
    week_rows, month_row = month_refs[m]
    t = f"'Tháng {m}'"
    cc = ws.cell(row=r, column=1, value=f"Tháng {m}")
    cc.font = f(bold=True, white=True)
    cc.fill = PatternFill("solid", fgColor=TAB_COLORS[m - 1])
    cc.alignment = Alignment(horizontal="center")
    cc.border = BORDER
    for k in range(5):
        cc = ws.cell(row=r, column=2 + k)
        if k < len(week_rows):
            cc.value = f"={t}!F{week_rows[k]}"
        cc.font = f(); cc.border = BORDER; cc.number_format = VND
        if m % 2 == 0:
            cc.fill = BAND_FILL
    cc = ws.cell(row=r, column=7, value=f"=SUM(B{r}:F{r})")
    cc.font = f(bold=True); cc.border = BORDER; cc.number_format = VND

r = 17
ws.row_dimensions[r].height = 24
cc = ws.cell(row=r, column=1, value="💰 CẢ NĂM")
cc.font = f(bold=True, size=12, white=True)
cc.alignment = Alignment(horizontal="center", vertical="center")
for j in range(1, 8):
    col = get_column_letter(j)
    cell = ws.cell(row=r, column=j)
    if j > 1:
        cell.value = f"=SUM({col}5:{col}16)"
        cell.number_format = VND
        cell.font = f(bold=True, size=12, white=True)
    cell.fill = GRAND_FILL
    cell.border = Border(top=med, bottom=med, left=thin, right=thin)

out = "/tmp/claude-0/-home-user-lien-ket-tiktok/26731859-2296-588b-8705-3fe299e2bd5a/scratchpad/Chi-phi-mua-hang-theo-ngay-2026.xlsx"
wb.save(out)

# ---- kiem tra ----
wb2 = load_workbook(out)
t1 = wb2["Tháng 1"]
th = wb2["TỔNG HỢP NĂM"]
w1, mrow1 = month_refs[1]
checks = [
    ("F4 (dong 1 ngay 1)", t1["F4"].value, "=D4*E4"),
    ("F12 (tong ngay 1)", t1["F12"].value, "=SUM(F4:F11)"),
    ("tuan 1 T1", t1[f"F{w1[0]}"].value, "=" + "+".join(f"F{x}" for x in [12, 21, 30, 39, 48, 57, 66])),
    ("tong thang 1", t1[f"F{mrow1}"].value, "=" + "+".join(f"F{x}" for x in w1)),
    ("TH B5 (tuan1 thang1)", th["B5"].value, f"='Tháng 1'!F{w1[0]}"),
    ("TH G5", th["G5"].value, "=SUM(B5:F5)"),
    ("TH G17", th["G17"].value, "=SUM(G5:G16)"),
]
ok = True
for name, got, want in checks:
    status = "OK " if got == want else "FAIL"
    if got != want: ok = False
    print(f"{status} {name}: {got}  (mong doi: {want})" if got != want else f"{status} {name}: {got}")
# so tuan: thang 2 co 4 tuan (28 ngay), cac thang khac 5 tuan
for m in range(1, 13):
    wr, _ = month_refs[m]
    want = 4 if m == 2 else 5
    assert len(wr) == want, f"Thang {m} co {len(wr)} tuan (mong doi {want})!"
print("sheets:", wb2.sheetnames)
print("ALL OK" if ok else "CO LOI!")
print("size:", os.path.getsize(out), "bytes")
