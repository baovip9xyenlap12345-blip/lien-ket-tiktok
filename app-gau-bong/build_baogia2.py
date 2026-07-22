#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""BAO GIA GAU BONG v2 — toi uu ban hang:
- Chia nhom danh muc co tieu de mau, size xep nho->lon, anh gop doc theo mau
- Cot Tinh trang (Con hang/Dat truoc), CF: dong da chon to xanh, vuot ton kho to cam
- Khoi thong tin khach hang, chan nhap sai so luong"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import os, re
from data_gau import PRODUCTS

ARIAL = "Arial"
PINK = "C2185B"
PINK2 = "E91E63"
PINK_LIGHT = "FCE4EC"
NAVY = "1F3864"
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
BAND_FILL = PatternFill("solid", fgColor="FDF2F6")
TOTAL_FILL = PatternFill("solid", fgColor="C2185B")
GREEN_ROW = PatternFill("solid", fgColor="D9EAD3")
ORANGE_WARN = PatternFill("solid", fgColor="F9CB9C")
thin = Side(style="thin", color="D0A5B8")
med = Side(style="medium", color="C2185B")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
VND = '#,##0'

def f(bold=False, size=10, white=False, blue=False, color="000000", italic=False):
    c = "FFFFFF" if white else ("0D47A1" if blue else color)
    return Font(name=ARIAL, bold=bold, size=size, color=c, italic=italic)

def base_name(t):
    return t.split(" (")[0]

def size_key(s):
    s = s.lower().replace(' ', '')
    m = re.match(r'^(\d)m(\d)?$', s)
    if m:
        return int(m.group(1)) * 100 + (int(m.group(2)) * 10 if m.group(2) else 0)
    m = re.match(r'^(\d+)', s.replace('s', ''))
    return int(m.group(1)) if m else 999

CATS = [
    ("🔥 GẤU HOT TREND", "E53935",
     ["Gấu Bông Capybara", "Capybara nằm", "Gấu bông labubu", "GẤU BÔNG CÁ SẤU Lông xù hot trend",
      "Gấu BabyThree", "Gấu bông Iphone 17", "Gấu bông dưa hấu", "Bơ dài", "Hoa hồng"]),
    ("🧸 GẤU BÔNG KINH ĐIỂN", "8E24AA",
     ["Gấu teddy", "Gấu dâu", "Gấu bắc cực", "Gấu spiderman", "Gấu bông stick", "Gấu bông lena 7 màu"]),
    ("🐰 THÚ BÔNG DỄ THƯƠNG", "039BE5",
     ["Thỏ bunny", "Thỏ tai dài", "Thỏ giận dữ lông xù qc", "Thỏ bình sữa", "Gấu Bông Thỏ Ôm Dâu",
      "Heo xám hông", "Heo ong vàng", "Heo bình sữa", "Hổ bình sữa", "Vịt VÀNG HỒNG VỊT TRẦM CẢM",
      "Vịt ôm bình sữa", "Chim cánh cụt", "Khủng long", "Mèo hoàng thượng QC", "Gấu Bông Mèo Ngáo Alaska",
      "Gấu Bông Hươu cao cổ", "Chó Ngáo"]),
    ("🎁 QUÀ TẶNG & GỐI", "43A047", ["Gối trái tim in logo"]),
]

# gom san pham theo base name, size nho -> lon
groups = {}
order = []
for p in PRODUCTS:
    b = base_name(p[0])
    if b not in groups:
        groups[b] = []
        order.append(b)
    groups[b].append(p)
for b in groups:
    groups[b].sort(key=lambda p: size_key(p[2]))

# danh sach cuoi cung: (category rows)
assigned = set()
final = []   # list of (cat_title, cat_color, [(ten,mau,size,gsi,gle,ghichu), ...group...])
for title, color, names in CATS:
    items = []
    for b in names:
        if b in groups:
            items.append((b, groups[b]))
            assigned.add(b)
    final.append((title, color, items))
missing = [b for b in order if b not in assigned]
assert not missing, f"SP chua gan danh muc: {missing}"

wb = Workbook()

# ================= TAB 1: BAO GIA & DAT HANG =================
ws = wb.active
ws.title = "BÁO GIÁ & ĐẶT HÀNG"
ws.sheet_properties.tabColor = PINK
ws.sheet_view.showGridLines = False
widths = {"A": 5, "B": 24, "C": 30, "D": 24, "E": 9, "F": 13, "G": 14, "H": 9, "I": 13, "J": 12, "K": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A6"

ws.merge_cells("A1:K1")
c = ws["A1"]
c.value = "🧸 XƯỞNG GẤU BÔNG BẢO BẢO — baogaubong.vn"
c.font = f(bold=True, size=18, white=True)
c.fill = PatternFill("solid", fgColor=PINK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34
ws.merge_cells("A2:K2")
c = ws["A2"]
c.value = "📞 0876.123.333 (call/Zalo)   ·   📍 Xưởng gấu Bảo Bảo, Phủ Yên 3, Yên Lập, Vĩnh Tường, Vĩnh Phúc"
c.font = f(bold=True, size=11, white=True)
c.fill = PatternFill("solid", fgColor=PINK2)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22
ws.merge_cells("A3:K3")
c = ws["A3"]
c.value = "Quý khách nhập SỐ LƯỢNG vào cột vàng → tiền tự tính · Dòng đã chọn tự tô XANH ✅ · Đặt nhiều hơn tồn kho dòng tô CAM (xưởng sản xuất thêm, gọi trước giúp xưởng nhé)"
c.font = f(size=9, italic=True, bold=True, color=PINK)
c.fill = PatternFill("solid", fgColor=PINK_LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[3].height = 24

headers = ["STT", "Ảnh sản phẩm", "Tên sản phẩm", "Màu sắc", "Size",
           "Giá lẻ\ntham khảo", "GIÁ SỈ\n(VNĐ/con)", "Tồn kho\n(con)", "Tình trạng",
           "SỐ LƯỢNG\nĐẶT ✍", "THÀNH TIỀN\n(VNĐ)"]
for j, lab in enumerate(headers, start=1):
    cc = ws.cell(row=5, column=j, value=lab)
    cc.font = f(bold=True, white=True, size=9)
    cc.fill = PatternFill("solid", fgColor=NAVY)
    cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cc.border = BORDER
ws.row_dimensions[5].height = 30

# ---- ghi du lieu ----
kho_list = []      # (ten, size, ghichu) theo dung thu tu kho
bg_of_kho = []     # dong bao gia ung voi tung dong kho
r = 6
stt = 0
gidx = 0
for title, color, items in final:
    ws.merge_cells(f"A{r}:K{r}")
    cc = ws.cell(row=r, column=1, value=f"{title}  ({sum(len(g) for _, g in items)} lựa chọn)")
    cc.font = f(bold=True, size=11, white=True)
    cc.fill = PatternFill("solid", fgColor=color)
    cc.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1
    for b, plist in items:
        gidx += 1
        band = gidx % 2 == 0
        first_r = r
        for (ten, mau, size, gsi, gle, ghichu) in plist:
            stt += 1
            kho_r = 4 + len(kho_list)
            kho_list.append((ten, size, ghichu))
            bg_of_kho.append(r)
            ws.row_dimensions[r].height = 24
            fill = BAND_FILL if band else None

            def cell(col, val, font=None, fmt=None, center=False):
                cc = ws.cell(row=r, column=col)
                if val is not None:
                    cc.value = val
                cc.font = font or f(size=9)
                cc.border = BORDER
                if fmt:
                    cc.number_format = fmt
                cc.alignment = Alignment(horizontal="center" if center else "left",
                                         vertical="center", wrap_text=not center)
                if fill:
                    cc.fill = fill
                return cc

            cell(1, stt, center=True)
            cell(2, None, center=True)
            cell(3, ten, font=f(size=9))
            cell(4, mau if mau else None, font=f(size=8))
            cell(5, size, center=True)
            cell(6, gle if gle > 0 else None, fmt=VND)
            cell(7, gsi, font=f(size=10, bold=True, color=PINK), fmt=VND)
            cell(8, f"='KHO HÀNG'!G{kho_r}", fmt='0', center=True)
            tc = cell(9, f'=IF(H{r}>0,"✅ Còn hàng","☎ Đặt trước")', font=f(size=8), center=True)
            sl = ws.cell(row=r, column=10)
            sl.font = f(bold=True, size=10, blue=True)
            sl.fill = INPUT_FILL
            sl.border = Border(left=med, right=med, top=thin, bottom=thin)
            sl.number_format = '0'
            sl.alignment = Alignment(horizontal="center", vertical="center")
            tt = ws.cell(row=r, column=11, value=f"=J{r}*G{r}")
            tt.font = f(bold=True, size=10)
            tt.border = BORDER
            tt.number_format = VND
            tt.alignment = Alignment(horizontal="right", vertical="center")
            if fill:
                tt.fill = fill
            r += 1
        # anh gop doc cho ca mau; mau chi co 1 size thi nang dong cao len de anh van to
        if r - 1 > first_r:
            ws.merge_cells(f"B{first_r}:B{r-1}")
        else:
            ws.row_dimensions[first_r].height = 55
        ic = ws.cell(row=first_r, column=2, value="🖼")
        ic.font = Font(name=ARIAL, size=16, color="C0C0C0")
        ic.alignment = Alignment(horizontal="center", vertical="center")

last_data = r - 1
N = len(kho_list)

# CF: dong da chon to xanh (A..K khi J>0); SL vuot ton kho to cam
rng = f"A6:K{last_data}"
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND($J6>0,$J6<=$H6)'], fill=GREEN_ROW))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND($J6>0,$J6>$H6)'], fill=ORANGE_WARN))
# validation so luong: so nguyen >= 0
dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True,
                    errorTitle="Số lượng không hợp lệ", error="Vui lòng nhập số nguyên ≥ 0",
                    showErrorMessage=True)
ws.add_data_validation(dv)
dv.add(f"J6:J{last_data}")

# ---- tong ----
r1 = last_data + 1
ws.merge_cells(f"A{r1}:I{r1}")
tc = ws.cell(row=r1, column=1, value="🧸 TỔNG SỐ GẤU ĐẶT (con)")
tc.font = f(bold=True, size=12, white=True)
tc.alignment = Alignment(horizontal="right", vertical="center")
for col in range(1, 12):
    ws.cell(row=r1, column=col).fill = TOTAL_FILL
    ws.cell(row=r1, column=col).border = Border(top=med, bottom=thin)
sc = ws.cell(row=r1, column=10, value=f"=SUM(J6:J{last_data})")
sc.font = f(bold=True, size=12, white=True); sc.number_format = '0'
sc.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r1].height = 24

r2 = r1 + 1
ws.merge_cells(f"A{r2}:I{r2}")
tc = ws.cell(row=r2, column=1, value="💰 TỔNG TIỀN QUÝ KHÁCH THANH TOÁN (VNĐ)")
tc.font = f(bold=True, size=13, white=True)
tc.alignment = Alignment(horizontal="right", vertical="center")
for col in range(1, 12):
    ws.cell(row=r2, column=col).fill = TOTAL_FILL
    ws.cell(row=r2, column=col).border = Border(top=thin, bottom=med)
ws.merge_cells(f"J{r2}:K{r2}")
sc = ws.cell(row=r2, column=10, value=f"=SUM(K6:K{last_data})")
sc.font = f(bold=True, size=14, white=True); sc.number_format = VND
sc.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r2].height = 28

# ---- thong tin khach hang ----
ci = r2 + 2
ws.merge_cells(f"A{ci}:K{ci}")
cc = ws.cell(row=ci, column=1, value="📋 THÔNG TIN KHÁCH HÀNG — quý khách điền giúp xưởng để giao hàng nhanh nhất")
cc.font = f(bold=True, size=11, white=True)
cc.fill = PatternFill("solid", fgColor=NAVY)
cc.alignment = Alignment(vertical="center")
ws.row_dimensions[ci].height = 22
ci_fields = ["Tên khách hàng / Shop:", "Số điện thoại / Zalo:", "Địa chỉ nhận hàng:", "Ghi chú (mẫu, màu, thời gian cần):"]
for k, lab in enumerate(ci_fields):
    rr = ci + 1 + k
    ws.merge_cells(f"A{rr}:C{rr}")
    lc = ws.cell(row=rr, column=1, value=lab)
    lc.font = f(bold=True, size=10)
    lc.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"D{rr}:K{rr}")
    ic = ws.cell(row=rr, column=4)
    ic.fill = INPUT_FILL
    ic.font = f(size=10, blue=True)
    ic.border = BORDER
    ws.row_dimensions[rr].height = 20

# ---- footer ----
foot = [
    "CÁCH ĐẶT HÀNG: ① Nhập số lượng vào cột vàng → tiền tự tính  ② Điền thông tin nhận hàng ở trên",
    "③ Gửi lại file này (hoặc chụp màn hình) qua Zalo 0876.123.333 — xưởng xác nhận và giao ngay.",
    "• '☎ Đặt trước': mẫu tạm hết tại xưởng nhưng vẫn nhận đặt — xưởng sản xuất trực tiếp, gọi để hẹn ngày lấy hàng.",
    "• Báo giá có hiệu lực 30 ngày kể từ ngày gửi. Giá sỉ áp dụng cho đơn số lượng — đơn lớn liên hệ để có giá tốt hơn.",
    "🧸 XƯỞNG GẤU BÔNG BẢO BẢO — Phủ Yên 3, Yên Lập, Vĩnh Tường, Vĩnh Phúc · 📞 0876.123.333 · 🌐 baogaubong.vn",
]
fr = ci + len(ci_fields) + 2
for k, text in enumerate(foot):
    rr = fr + k
    ws.merge_cells(f"A{rr}:K{rr}")
    cc = ws.cell(row=rr, column=1, value=text)
    cc.font = f(bold=(k == 4), size=10, color=PINK if k == 4 else "000000")

# ================= TAB 2: KHO HANG =================
ws = wb.create_sheet("KHO HÀNG")
ws.sheet_properties.tabColor = "757575"
ws.sheet_view.showGridLines = False
for col, w in {"A": 5, "B": 32, "C": 10, "D": 12, "E": 12, "F": 12, "G": 14, "H": 22}.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A4"
ws.merge_cells("A1:H1")
c = ws["A1"]
c.value = "KHO HÀNG THỰC TẾ — CHỦ XƯỞNG QUẢN LÝ (ẨN TAB NÀY TRƯỚC KHI GỬI KHÁCH)"
c.font = f(bold=True, size=13, white=True)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26
ws.merge_cells("A2:H2")
c = ws["A2"]
c.value = "Nhập 3 cột vàng: Tồn đầu · Nhập thêm · Đã bán → 'Tồn hiện tại' tự tính và tự hiện sang tab BÁO GIÁ (cột Tồn kho + Tình trạng)"
c.font = f(size=9, italic=True, color="808080")
c.alignment = Alignment(horizontal="center")

for j, lab in enumerate(["STT", "Tên sản phẩm", "Size", "Tồn đầu ✍", "Nhập thêm ✍", "Đã bán ✍",
                          "Tồn hiện tại", "Ghi chú (từ bảng giá gốc)"], start=1):
    cc = ws.cell(row=3, column=j, value=lab)
    cc.font = f(bold=True, white=True, size=9)
    cc.fill = PatternFill("solid", fgColor=NAVY)
    cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cc.border = BORDER

for i, (ten, size, ghichu) in enumerate(kho_list):
    r = 4 + i
    for j in range(1, 9):
        ws.cell(row=r, column=j).border = BORDER
    ws.cell(row=r, column=1, value=i + 1).font = f(size=9)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2, value=ten).font = f(size=9)
    ws.cell(row=r, column=3, value=size).font = f(size=9)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    for j in (4, 5, 6):
        cc = ws.cell(row=r, column=j)
        cc.font = f(size=9, blue=True)
        cc.fill = INPUT_FILL
        cc.number_format = '0'
        cc.alignment = Alignment(horizontal="center")
    ton = ws.cell(row=r, column=7, value=f"=D{r}+E{r}-F{r}")
    ton.font = f(size=9, bold=True)
    ton.number_format = '0'
    ton.alignment = Alignment(horizontal="center")
    if ghichu:
        gc = ws.cell(row=r, column=8, value=ghichu)
        gc.font = f(size=8, italic=True, color="808080")

# ================= TAB 3: HUONG DAN =================
ws = wb.create_sheet("HƯỚNG DẪN")
ws.sheet_properties.tabColor = "FFB300"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 112
guide = [
    ("HƯỚNG DẪN SỬ DỤNG — FILE BÁO GIÁ GẤU BÔNG BẢO BẢO (bản tối ưu bán hàng)", 1),
    ("", 0),
    ("ĐIỂM MỚI CỦA BẢN NÀY:", 2),
    ("  • Sản phẩm chia 4 nhóm màu: 🔥 Hot Trend · 🧸 Gấu kinh điển · 🐰 Thú bông dễ thương · 🎁 Quà tặng & gối — khách lướt tìm nhanh.", 0),
    ("  • Size xếp từ nhỏ → lớn trong từng mẫu. Ô ảnh gộp dọc: mỗi mẫu chỉ cần chèn 1 ảnh đại diện.", 0),
    ("  • Cột 'Tình trạng' tự đổi: còn kho → '✅ Còn hàng'; hết kho → '☎ Đặt trước' (vẫn nhận đơn, không mất khách).", 0),
    ("  • Khách nhập số lượng: dòng tự tô XANH LÁ = đã chọn hợp lệ; tô CAM = đặt vượt tồn kho (xưởng cần sản xuất thêm).", 0),
    ("  • Cột số lượng chỉ nhận số nguyên ≥ 0 — khách gõ nhầm chữ sẽ bị báo lỗi ngay.", 0),
    ("  • Cuối trang có khối THÔNG TIN KHÁCH HÀNG (tên, SĐT, địa chỉ, ghi chú) → nhận file lại là đủ thông tin lên đơn luôn.", 0),
    ("", 0),
    ("VIỆC CỦA CHỦ XƯỞNG:", 2),
    ("  1. CHÈN ẢNH: click ô 🖼 của từng mẫu → Chèn (Insert) → Hình ảnh → 'Chèn hình ảnh trong ô'. Mỗi mẫu 1 ảnh (ô đã gộp dọc).", 0),
    ("  2. CẬP NHẬT KHO: tab 'KHO HÀNG' — điền Tồn đầu / Nhập thêm / Đã bán → tồn kho + tình trạng bên báo giá tự đổi.", 0),
    ("  3. TRƯỚC KHI GỬI KHÁCH: chuột phải tab 'KHO HÀNG' và 'HƯỚNG DẪN' → Ẩn trang tính (Hide sheet).", 0),
    ("  4. GỬI KHÁCH: Tệp → Tạo bản sao cho từng khách (khách tự nhập số lượng vào bản của họ), hoặc chia sẻ link 'Người xem'.", 0),
    ("  5. ĐỔI GIÁ: sửa thẳng cột GIÁ SỈ. Thêm mẫu mới: chèn dòng vào đúng nhóm + thêm dòng tương ứng trong tab KHO HÀNG.", 0),
    ("  6. KHO Ở FILE RIÊNG (nếu muốn): thay công thức cột Tồn kho bằng =IMPORTRANGE(\"link-file-kho\"; \"KHO HÀNG!G4\")", 0),
    ("     rồi bấm 'Cho phép truy cập' lần đầu.", 0),
    ("", 0),
    ("Nguồn giá: 'Bảng Giá Sỉ SP baogaubong.vn' (tab GẤU BÔNG ALL), trích ngày 21/07/2026 — 63 mẫu/size có giá sỉ.", 0),
    ("Gợi ý: kiểm tra và thêm số tài khoản ngân hàng vào chân trang báo giá để khách chuyển khoản nhanh (sửa dòng cuối tab BÁO GIÁ).", 0),
]
for i, (text, style) in enumerate(guide, start=1):
    c = ws.cell(row=i, column=2, value=text)
    c.font = Font(name=ARIAL, bold=True, size=14, color=PINK) if style == 1 else f(bold=(style > 0), size=11)

out = "/tmp/claude-0/-home-user-lien-ket-tiktok/26731859-2296-588b-8705-3fe299e2bd5a/scratchpad/Bao-gia-gau-bong-Bao-Bao.xlsx"
wb.save(out)

# ---- kiem tra nguoc ----
wb2 = load_workbook(out)
bg = wb2["BÁO GIÁ & ĐẶT HÀNG"]
kho = wb2["KHO HÀNG"]
ok = True
def chk(name, got, want):
    global ok
    good = got == want
    if not good: ok = False
    print(("OK  " if good else "FAIL") + f" {name}: {got}" + ("" if good else f" (mong doi {want})"))

# 1) moi dong kho khop dong bao gia (ten + size + cong thuc ton)
mismatch = 0
for i, bgr in enumerate(bg_of_kho):
    kr = 4 + i
    if bg.cell(row=bgr, column=3).value != kho.cell(row=kr, column=2).value: mismatch += 1
    if bg.cell(row=bgr, column=5).value != kho.cell(row=kr, column=3).value: mismatch += 1
    if bg.cell(row=bgr, column=8).value != f"='KHO HÀNG'!G{kr}": mismatch += 1
chk("khop bao gia <-> kho (63 dong x3)", mismatch, 0)
chk("so dong kho", len(kho_list), 63)
# 2) cong thuc mau
r0 = bg_of_kho[0]
chk("thanh tien dong dau", bg.cell(row=r0, column=11).value, f"=J{r0}*G{r0}")
chk("tinh trang dong dau", bg.cell(row=r0, column=9).value, f'=IF(H{r0}>0,"✅ Còn hàng","☎ Đặt trước")')
chk("kho G4", kho["G4"].value, "=D4+E4-F4")
# 3) size da xep tang dan trong tung mau
bad_sort = 0
prev_b, prev_k = None, -1
for i, bgr in enumerate(bg_of_kho):
    b = base_name(bg.cell(row=bgr, column=3).value)
    k = size_key(bg.cell(row=bgr, column=5).value)
    if b == prev_b and k < prev_k: bad_sort += 1
    prev_b, prev_k = b, k
chk("size tang dan trong tung mau", bad_sort, 0)
# 4) gia doi chieu nguon
prices = {}
for bgr in bg_of_kho:
    prices[(bg.cell(row=bgr, column=3).value, bg.cell(row=bgr, column=5).value)] = bg.cell(row=bgr, column=7).value
chk("Capybara (Rút mũi) 70cm", prices.get(("Gấu Bông Capybara (Rút mũi)", "70cm")), 150000)
chk("Teddy 1m2", prices.get(("Gấu teddy", "1m2")), 190000)
chk("Ca sau 2m", prices.get(("GẤU BÔNG CÁ SẤU Lông xù hot trend", "2m")), 300000)
chk("Labubu 70cm", prices.get(("Gấu bông labubu", "70cm")), 80000)
chk("Bac cuc 1m", prices.get(("Gấu bắc cực", "1m")), 400000)
# 5) CF & validation ton tai
chk("so rule CF", len(bg.conditional_formatting), 1)  # 1 range voi 2 rule
chk("so validation", len(bg.data_validations.dataValidation), 1)
print("tabs:", wb2.sheetnames)
print("ALL OK" if ok else "CO LOI!")
print("size:", os.path.getsize(out), "bytes")
