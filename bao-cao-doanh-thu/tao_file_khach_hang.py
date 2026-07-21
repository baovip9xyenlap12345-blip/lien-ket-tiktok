# -*- coding: utf-8 -*-
"""File 'thong tin khach hang dat hang': tu keo du lieu khach tu file doanh thu
qua IMPORTRANGE + QUERY (chi chay tren Google Sheets — day la dich den cua file)."""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MASTER_URL = "https://docs.google.com/spreadsheets/d/1EK5edM1bzp-A5_BLlZn06DeqVvFG0xaib3u0vkpxnfQ"
MONTHS = [6, 7, 8, 9, 10, 11, 12]
OUT = "/tmp/claude-0/-home-user-lien-ket-tiktok/cddf1fe0-bb70-57dc-ae93-5dfdde080200/scratchpad/thong tin khach hang dat hang.xlsx"

FONT = "Arial"
f_title = Font(name=FONT, size=14, bold=True, color="FFFFFF")
f_head = Font(name=FONT, size=10, bold=True, color="FFFFFF")
f_bold = Font(name=FONT, size=11, bold=True)
f_norm = Font(name=FONT, size=10)
f_blue = Font(name=FONT, size=10, color="0000FF")
f_note = Font(name=FONT, size=10, italic=True, color="808080")
fill_title = PatternFill("solid", fgColor="1F4E78")
fill_head = PatternFill("solid", fgColor="2E75B6")
fill_yellow = PatternFill("solid", fgColor="FFFF00")
thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ------------------------------------------------ 1. KHACH DAT HANG
ws = wb.active
ws.title = "KHÁCH ĐẶT HÀNG"
for col, w in {"A": 12, "B": 22, "C": 14, "D": 30, "E": 24, "F": 22, "G": 15, "H": 18}.items():
    ws.column_dimensions[col].width = w
ws.column_dimensions["A"].number_format = "dd/mm/yyyy"
ws.column_dimensions["G"].number_format = "#,##0"
for cl in "ABCDEFH":
    ws.column_dimensions[cl].font = f_norm

ws.merge_cells("A1:H1")
c = ws["A1"]; c.value = "DANH SÁCH KHÁCH ĐẶT HÀNG — TỰ CẬP NHẬT TỪ FILE BÁO CÁO DOANH THU"
c.font = f_title; c.fill = fill_title; c.alignment = center
ws.row_dimensions[1].height = 26

ws["A3"] = "Link file doanh thu (nguồn):"
ws["A3"].font = f_bold
ws.merge_cells("B3:E3")
b3 = ws["B3"]; b3.value = MASTER_URL
b3.font = f_blue; b3.fill = fill_yellow; b3.alignment = left; b3.border = border
ws["F3"] = "Trạng thái kết nối →"
ws["F3"].font = f_bold
# o kich hoat ket noi: bam vao o G3 va bam "Cho phep truy cap" (Allow access)
g3 = ws["G3"]
g3.value = "=IF(ISERROR(IMPORTRANGE($B$3,\"'Tháng 6'!A1\")),\"BẤM VÀO ĐÂY → Cho phép truy cập\",\"✔ ĐÃ KẾT NỐI\")"
g3.font = f_bold; g3.fill = fill_yellow; g3.border = border; g3.alignment = center
ws.merge_cells("G3:H3")

headers = ["NGÀY ĐẶT", "TÊN KHÁCH HÀNG", "SỐ ĐIỆN THOẠI", "ĐỊA CHỈ GIAO HÀNG",
           "CÔNG TY", "SẢN PHẨM ĐÃ ĐẶT", "THÀNH TIỀN (đ)", "CHUYỂN KHOẢN"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=5, column=i, value=h)
    c.font = f_head; c.fill = fill_head; c.alignment = center; c.border = border
ws.row_dimensions[5].height = 28

parts = ";".join(f"IMPORTRANGE($B$3,\"'Tháng {m}'!B8:N407\")" for m in MONTHS)
ws["A6"] = ("=IFERROR(QUERY({" + parts + "},"
            "\"select Col1,Col2,Col5,Col12,Col3,Col6,Col9,Col11 "
            "where Col1 is not null order by Col1\",0),"
            "\"⚠ Chưa kết nối: bấm vào ô màu vàng ở dòng 3 rồi bấm nút Cho phép truy cập (Allow access)\")")
ws["A6"].font = f_norm
ws.freeze_panes = "A6"

# ------------------------------------------------ 2. TONG HOP KHACH
ws2 = wb.create_sheet("TỔNG HỢP KHÁCH")
for col, w in {"A": 22, "B": 14, "C": 30, "D": 10, "E": 16}.items():
    ws2.column_dimensions[col].width = w
ws2.column_dimensions["E"].number_format = "#,##0"
ws2.merge_cells("A1:E1")
c = ws2["A1"]; c.value = "TỔNG HỢP THEO KHÁCH HÀNG (số đơn & tổng tiền, tự gộp)"
c.font = f_title; c.fill = fill_title; c.alignment = center
ws2.row_dimensions[1].height = 26
for i, h in enumerate(["TÊN KHÁCH HÀNG", "SỐ ĐIỆN THOẠI", "ĐỊA CHỈ", "SỐ ĐƠN", "TỔNG TIỀN (đ)"], start=1):
    c = ws2.cell(row=3, column=i, value=h)
    c.font = f_head; c.fill = fill_head; c.alignment = center; c.border = border
ws2["A4"] = ("=IFERROR(QUERY('KHÁCH ĐẶT HÀNG'!A6:H,"
             "\"select B,C,D,count(A),sum(G) where B is not null "
             "group by B,C,D order by sum(G) desc label count(A) '', sum(G) ''\",0),"
             "\"Chưa có dữ liệu — hãy kết nối ở sheet KHÁCH ĐẶT HÀNG trước\")")
ws2["A4"].font = f_norm
ws2.freeze_panes = "A4"

# ------------------------------------------------ 3. HUONG DAN
ws3 = wb.create_sheet("HƯỚNG DẪN")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 105
ws3["B2"] = "HƯỚNG DẪN — FILE THÔNG TIN KHÁCH HÀNG (TỰ ĐẤU NỐI VỚI FILE BÁO CÁO DOANH THU)"
ws3["B2"].font = Font(name=FONT, size=13, bold=True, color="1F4E78")
lines = [
    "1. KẾT NỐI LẦN ĐẦU (chỉ làm 1 lần):",
    "   • Mở sheet 'KHÁCH ĐẶT HÀNG' → bấm vào ô màu VÀNG ở dòng 3 (bên phải chữ 'Trạng thái kết nối').",
    "   • Bấm nút 'Cho phép truy cập' (Allow access) hiện ra → vài giây sau dữ liệu tự đổ về, ô chuyển thành ✔ ĐÃ KẾT NỐI.",
    "",
    "2. CÁCH HOẠT ĐỘNG:",
    "   • Bạn CHỈ nhập đơn hàng ở file 'doanh thu thông tin đơn hàng' như bình thường.",
    "   • File này TỰ ĐỘNG kéo sang: Ngày đặt · Tên khách · SĐT · Địa chỉ giao hàng · Công ty · Sản phẩm · Thành tiền · Tình trạng CK.",
    "   • KHÔNG gõ tay vào vùng dữ liệu của file này — mọi thứ tự cập nhật, gõ đè sẽ làm hỏng công thức.",
    "",
    "3. SHEET 'TỔNG HỢP KHÁCH': tự gộp mỗi khách 1 dòng, đếm số đơn và cộng tổng tiền, xếp khách mua nhiều lên đầu.",
    "",
    "4. NẾU ĐỔI FILE DOANH THU SANG LINK KHÁC: dán link mới vào ô vàng B3 của sheet 'KHÁCH ĐẶT HÀNG' rồi bấm Cho phép truy cập lại.",
    "",
    "Lưu ý: file doanh thu nguồn phải là Google Sheet 'xịn' (link bắt đầu docs.google.com/spreadsheets và KHÔNG có đuôi .xlsx trên tiêu đề).",
]
r = 4
for t in lines:
    ws3.cell(row=r, column=2, value=t).font = f_bold if t and t[0].isdigit() else f_norm
    ws3.cell(row=r, column=2).alignment = left
    r += 1

wb.save(OUT)
print("Saved:", OUT)
