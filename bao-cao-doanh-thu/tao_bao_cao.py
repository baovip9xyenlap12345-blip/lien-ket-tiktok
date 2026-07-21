# -*- coding: utf-8 -*-
"""File 'doanh thu thong tin don hang.xlsx' — bo cuc 4 TUAN / thang:
- Sheet HƯỚNG DẪN
- Sheet Tháng 6..Tháng 12 (2026): moi thang chia 4 khoi TUAN (1-7, 8-14,
  15-21, 22-cuoi thang), moi tuan 80 dong nhap + banner doanh thu tuan;
  ben phai la bang doanh thu theo ngay
- Sheet BÁO CÁO NĂM: tra cuu ngay, bang tuan theo thang, thang, quy, nam
"""
import datetime as dt
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

YEAR = 2026
MONTHS = [6, 7, 8, 9, 10, 11, 12]
WEEK_ROWS = 80              # so dong nhap moi tuan (~10 don/ngay van du)
BLOCK = WEEK_ROWS + 2       # banner + header + du lieu
FULL_FIRST = 10             # dong du lieu dau tien (tuan 1)
FULL_LAST = 8 + 4 * BLOCK - 1   # 335 — dong du lieu cuoi (tuan 4)
OUT = "/tmp/claude-0/-home-user-lien-ket-tiktok/cddf1fe0-bb70-57dc-ae93-5dfdde080200/scratchpad/doanh thu thong tin don hang.xlsx"

FONT = "Arial"
VND = '#,##0;(#,##0);"-"'
DATE_FMT = "dd/mm/yyyy"

f_title = Font(name=FONT, size=14, bold=True, color="FFFFFF")
f_head = Font(name=FONT, size=10, bold=True, color="FFFFFF")
f_banner = Font(name=FONT, size=11, bold=True, color="FFFFFF")
f_bold = Font(name=FONT, size=11, bold=True)
f_norm = Font(name=FONT, size=10)
f_blue = Font(name=FONT, size=10, color="0000FF")
f_green = Font(name=FONT, size=10, color="008000")
f_note = Font(name=FONT, size=10, italic=True, color="808080")

fill_title = PatternFill("solid", fgColor="1F4E78")
fill_head = PatternFill("solid", fgColor="2E75B6")
fill_sub = PatternFill("solid", fgColor="DDEBF7")
fill_yellow = PatternFill("solid", fgColor="FFFF00")
fill_green = PatternFill("solid", fgColor="C6EFCE")
fill_amber = PatternFill("solid", fgColor="FFEB9C")   # da coc
fill_red = PatternFill("solid", fgColor="FFC7CE")
BANNER_FILLS = [PatternFill("solid", fgColor=c)
                for c in ("4472C4", "305496", "4472C4", "305496")]
TAB_COLORS = {6: "70AD47", 7: "4472C4", 8: "ED7D31", 9: "FFC000",
              10: "7030A0", 11: "00B0B0", 12: "C00000"}

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

def week_ranges(m):
    last = calendar.monthrange(YEAR, m)[1]
    return [(1, 7), (8, 14), (15, 21), (22, last)]

def block_rows(w):
    """w = 1..4 -> (dong banner, dong header, dong dau, dong cuoi)"""
    bs = 8 + (w - 1) * BLOCK
    return bs, bs + 1, bs + 2, bs + 1 + WEEK_ROWS

wb = Workbook()

# ---------------------------------------------------------------- HUONG DAN
ws = wb.active
ws.title = "HƯỚNG DẪN"
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = "808080"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 110
ws["B2"] = "HƯỚNG DẪN SỬ DỤNG FILE BÁO CÁO DOANH THU - THÔNG TIN ĐƠN HÀNG (THÁNG 6 - THÁNG 12/2026)"
ws["B2"].font = Font(name=FONT, size=14, bold=True, color="1F4E78")

lines = [
    ("1. CẤU TRÚC FILE", True),
    ("   • 7 sheet 'Tháng 6' → 'Tháng 12': nhập đơn hàng của từng tháng, mỗi tháng 1 sheet riêng (mỗi tab 1 màu).", False),
    ("   • Mỗi tháng chia sẵn 4 KHỐI TUẦN: Tuần 1 (ngày 1-7) · Tuần 2 (ngày 8-14) · Tuần 3 (ngày 15-21) · Tuần 4 (ngày 22-cuối tháng).", False),
    ("   • Mỗi tuần có 80 dòng nhập; ngay trên đầu mỗi khối là DOANH THU TUẦN tự cộng.", False),
    ("   • Sheet 'BÁO CÁO NĂM': tự động tính doanh thu theo NGÀY - TUẦN - THÁNG - QUÝ - NĂM, không cần nhập gì thêm.", False),
    ("", False),
    ("2. CÁCH NHẬP LIỆU (trong các sheet tháng)", True),
    ("   • Nhập đơn vào đúng KHỐI TUẦN chứa ngày đặt (ví dụ đơn ngày 21/07 → khối TUẦN 3 của sheet Tháng 7).", False),
    ("   • Chỉ nhập các cột: NGÀY ĐẶT, TÊN KHÁCH HÀNG, CÔNG TY, MÃ SỐ THUẾ, SỐ ĐIỆN THOẠI, TÊN SẢN PHẨM, SỐ LƯỢNG, ĐƠN GIÁ, HÌNH ẢNH, TÌNH TRẠNG CHUYỂN KHOẢN, ĐỊA CHỈ GIAO HÀNG, GHI CHÚ.", False),
    ("   • Cột STT và THÀNH TIỀN tự tính, KHÔNG sửa (Thành tiền = Số lượng × Đơn giá).", False),
    ("   • NGÀY ĐẶT nhập dạng dd/mm/yyyy (ví dụ 05/07/2026) — bắt buộc đúng định dạng thì báo cáo mới tính được.", False),
    ("   • SỐ LƯỢNG và ĐƠN GIÁ chỉ gõ số (ví dụ 50 và 85000) — không gõ chữ 'cái', 'đ', dấu chấm phẩy.", False),
    ("   • TÌNH TRẠNG CHUYỂN KHOẢN: bấm vào ô sẽ có danh sách chọn 'Đã chuyển khoản' / 'Đã cọc' / 'Chưa chuyển khoản' (ô tự đổi màu XANH / VÀNG / ĐỎ).", False),
    ("   • Khách đã cọc: chọn 'Đã cọc' — khi khách thanh toán đủ thì đổi lại thành 'Đã chuyển khoản'. Tổng tiền các đơn đã cọc hiện ở đầu sheet và trong BÁO CÁO NĂM.", False),
    ("   • Sheet 'Tháng 6' có sẵn 3 dòng VÍ DỤ MẪU (chữ màu xanh dương) — hãy xóa hoặc ghi đè bằng dữ liệu thật của bạn.", False),
    ("", False),
    ("3. CÁCH CHÈN HÌNH ẢNH SẢN PHẨM VÀO Ô", True),
    ("   • Trên Google Sheets: bấm chọn ô ở cột HÌNH ẢNH SẢN PHẨM → menu Chèn (Insert) → Hình ảnh (Image) → 'Chèn hình ảnh trong ô' (Insert image in cell).", False),
    ("   • Trên Excel: chọn ô → Insert → Pictures → Place in Cell.", False),
    ("", False),
    ("4. BẢNG DOANH THU THEO NGÀY (bên phải mỗi sheet tháng, cột P-Q)", True),
    ("   • Tự cộng doanh thu của từng ngày trong tháng đó, kể cả nhiều đơn cùng 1 ngày.", False),
    ("   • File 'thông tin khách hàng đặt hàng' (file riêng) tự kéo NGÀY ĐẶT - TÊN - SĐT - ĐỊA CHỈ từ file này sang, không cần nhập lại.", False),
    ("", False),
    ("5. SHEET 'BÁO CÁO NĂM'", True),
    ("   • TRA CỨU THEO NGÀY: gõ 1 ngày bất kỳ vào ô màu VÀNG → ô bên cạnh hiện doanh thu của đúng ngày đó.", False),
    ("   • DOANH THU THEO TUẦN: mỗi tháng 4 tuần (đúng theo khối tuần bạn nhập), tự cộng.", False),
    ("   • DOANH THU THEO THÁNG / QUÝ / NĂM: tự động lấy từ các sheet tháng, kèm tổng ĐÃ và CHƯA chuyển khoản.", False),
    ("", False),
    ("6. THÊM DÒNG KHI 1 TUẦN HẾT CHỖ", True),
    ("   • Bôi đen 1 dòng Ở GIỮA khối tuần đó → chuột phải → 'Chèn 1 hàng phía trên', rồi kéo công thức cột STT và THÀNH TIỀN từ dòng trên xuống.", False),
    ("   • KHÔNG nhập dữ liệu ra ngoài các khối tuần — báo cáo sẽ không cộng phần đó.", False),
    ("", False),
    ("7. QUY ƯỚC MÀU", True),
    ("   • Chữ XANH DƯƠNG: dữ liệu ví dụ / ô người dùng nhập.  • Chữ ĐEN: công thức tự tính.  • Chữ XANH LÁ: số lấy từ sheet khác.  • Ô nền VÀNG: ô cần bạn nhập.", False),
]
r = 4
for text, bold in lines:
    ws.cell(row=r, column=2, value=text).font = f_bold if bold else f_norm
    ws.cell(row=r, column=2).alignment = left
    r += 1

# ---------------------------------------------------------------- SHEET THANG
HEADERS = ["STT", "NGÀY ĐẶT", "TÊN KHÁCH HÀNG", "CÔNG TY", "MÃ SỐ THUẾ",
           "SỐ ĐIỆN THOẠI", "TÊN SẢN PHẨM", "SỐ LƯỢNG", "ĐƠN GIÁ (đ)",
           "THÀNH TIỀN (đ)", "HÌNH ẢNH SẢN PHẨM", "TÌNH TRẠNG CHUYỂN KHOẢN",
           "ĐỊA CHỈ GIAO HÀNG", "GHI CHÚ"]
WIDTHS = [5, 12, 20, 26, 14, 14, 24, 10, 12, 14, 18, 20, 26, 18]

EXAMPLES = [  # chi dat o sheet Thang 6 — vi du mau, thay bang du lieu that
    [dt.date(2026, 6, 3), "Nguyễn Văn An", "Công ty TNHH Thương Mại ABC", "0312345678",
     "0901234567", "Bình giữ nhiệt in logo", 100, 85000, "Đã chuyển khoản",
     "12 Nguyễn Trãi, Thanh Xuân, Hà Nội", "Giao đợt 1 (VÍ DỤ MẪU)"],
    [dt.date(2026, 6, 10), "Trần Thị Bích", "Công ty CP Đầu Tư XYZ", "0109876543",
     "0987654321", "Sổ tay da in logo", 50, 120000, "Chưa chuyển khoản",
     "45 Lê Lợi, Quận 1, TP.HCM", "(VÍ DỤ MẪU)"],
    [dt.date(2026, 6, 15), "Lê Minh Cường", "Công ty TNHH DEF", "0301122334",
     "0912345678", "Bút ký khắc tên", 200, 35000, "Đã chuyển khoản",
     "88 Trần Phú, Hải Châu, Đà Nẵng", "Khách quen (VÍ DỤ MẪU)"],
]

REAL_T7 = [  # 2 don that nguoi dung da nhap — giu nguyen (ngay 21/07 -> TUAN 3)
    [dt.date(2026, 7, 21), "cty flc", "ctyflc", "2500740944",
     "0866688324", "tedy 25cm", 1000, 50000, "Đã chuyển khoản", "", "khách cũ"],
    [dt.date(2026, 7, 21), "flc 1", "nvk", "",
     "0866688324", "tedty 30", 1000, 47000, "Chưa chuyển khoản", "", ""],
]

from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

def month_sheet(m):
    name = f"Tháng {m}"
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLORS[m]
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["O"].width = 3
    ws.column_dimensions["P"].width = 12
    ws.column_dimensions["Q"].width = 16

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"BÁO CÁO DOANH THU THÁNG {m}/{YEAR}"
    c.font = f_title; c.fill = fill_title; c.alignment = center
    ws.row_dimensions[1].height = 26

    # khoi tong hop thang
    labels = [
        ("TỔNG DOANH THU THÁNG (đ):", f"=SUM($J${FULL_FIRST}:$J${FULL_LAST})"),
        ("ĐÃ CHUYỂN KHOẢN (đ):", f'=SUMIF($L${FULL_FIRST}:$L${FULL_LAST},"Đã chuyển khoản",$J${FULL_FIRST}:$J${FULL_LAST})'),
        ("ĐÃ CỌC (đ):", f'=SUMIF($L${FULL_FIRST}:$L${FULL_LAST},"Đã cọc",$J${FULL_FIRST}:$J${FULL_LAST})'),
        ("CHƯA CHUYỂN KHOẢN (đ):", f'=SUMIF($L${FULL_FIRST}:$L${FULL_LAST},"Chưa chuyển khoản",$J${FULL_FIRST}:$J${FULL_LAST})'),
    ]
    for i, (lab, formula) in enumerate(labels):
        rr = 3 + i
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
        lc = ws.cell(row=rr, column=2, value=lab)
        lc.font = f_bold; lc.fill = fill_sub; lc.alignment = left
        vc = ws.cell(row=rr, column=4, value=formula)
        vc.font = f_bold; vc.number_format = VND; vc.alignment = right; vc.fill = fill_sub

    # dinh dang theo cot (o moi nguoi dung them van nhan dung dinh dang)
    col_styles = {
        "A": dict(font=f_norm, alignment=center),
        "B": dict(font=f_norm, number_format=DATE_FMT, alignment=center),
        "C": dict(font=f_norm), "D": dict(font=f_norm),
        "E": dict(font=f_norm, number_format="@"),
        "F": dict(font=f_norm, number_format="@"),
        "G": dict(font=f_norm),
        "H": dict(font=f_norm, number_format="#,##0", alignment=right),
        "I": dict(font=f_norm, number_format=VND, alignment=right),
        "J": dict(font=f_norm, number_format=VND, alignment=right),
        "K": dict(font=f_norm, alignment=center),
        "L": dict(font=f_norm, alignment=center),
        "M": dict(font=f_norm, alignment=left),
        "N": dict(font=f_norm),
    }
    for colL, props in col_styles.items():
        d = ws.column_dimensions[colL]
        for k, v in props.items():
            setattr(d, k, v)

    dv = DataValidation(type="list", formula1='"Đã chuyển khoản,Đã cọc,Chưa chuyển khoản"',
                        allow_blank=True, showDropDown=False)
    dv.error = "Chọn: Đã chuyển khoản / Đã cọc / Chưa chuyển khoản"
    ws.add_data_validation(dv)

    # 4 khoi tuan
    for w, (d1, d2) in enumerate(week_ranges(m), start=1):
        bs, hr, first, last = block_rows(w)
        # banner tuan
        ws.merge_cells(start_row=bs, start_column=1, end_row=bs, end_column=10)
        bc = ws.cell(row=bs, column=1,
                     value=f"TUẦN {w}  ({d1:02d}/{m:02d} → {d2:02d}/{m:02d})   —   DOANH THU TUẦN (đ):")
        bc.font = f_banner; bc.fill = BANNER_FILLS[w - 1]; bc.alignment = left
        ws.merge_cells(start_row=bs, start_column=11, end_row=bs, end_column=14)
        vc = ws.cell(row=bs, column=11, value=f"=SUM($J${first}:$J${last})")
        vc.font = f_banner; vc.fill = BANNER_FILLS[w - 1]
        vc.number_format = VND; vc.alignment = right
        ws.row_dimensions[bs].height = 22
        # header bang
        for i, h in enumerate(HEADERS, start=1):
            c = ws.cell(row=hr, column=i, value=h)
            c.font = f_head; c.fill = fill_head; c.alignment = center; c.border = border
        ws.row_dimensions[hr].height = 28
        # dong du lieu (ke khung + dinh dang tung o cho dep)
        for rr in range(first, last + 1):
            for col in range(1, 15):
                ws.cell(row=rr, column=col).border = border
            ac = ws.cell(row=rr, column=1, value=f'=IF($B{rr}="","",ROW()-{hr})')
            ac.font = f_norm; ac.alignment = center
            ws.cell(row=rr, column=2).number_format = DATE_FMT
            ws.cell(row=rr, column=5).number_format = "@"
            ws.cell(row=rr, column=6).number_format = "@"
            ws.cell(row=rr, column=8).number_format = "#,##0"
            ws.cell(row=rr, column=9).number_format = VND
            jc = ws.cell(row=rr, column=10, value=f'=IF($H{rr}="","",$H{rr}*$I{rr})')
            jc.font = f_norm; jc.number_format = VND; jc.alignment = right
        dv.add(f"L{first}:L{last}")

    # to mau tinh trang chuyen khoan (toan vung)
    rng = f"L{FULL_FIRST}:L{FULL_LAST}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal",
        formula=['"Đã chuyển khoản"'], fill=fill_green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal",
        formula=['"Đã cọc"'], fill=fill_amber))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal",
        formula=['"Chưa chuyển khoản"'], fill=fill_red))

    # vi du mau (Thang 6) / don that (Thang 7) — dat vao dung khoi tuan
    rows_to_fill = EXAMPLES if m == 6 else (REAL_T7 if m == 7 else [])
    nxt = {}
    for row in rows_to_fill:
        w = min((row[0].day - 1) // 7, 3) + 1
        bs, hr, first, last = block_rows(w)
        rr = nxt.get(w, first)
        nxt[w] = rr + 1
        vals = {2: row[0], 3: row[1], 4: row[2], 5: row[3], 6: row[4],
                7: row[5], 8: row[6], 9: row[7], 12: row[8], 13: row[9], 14: row[10]}
        for col, v in vals.items():
            if v == "":
                continue
            c = ws.cell(row=rr, column=col, value=v)
            c.font = f_blue
        if m == 6:
            ws.cell(row=rr, column=11, value="(chèn ảnh vào ô này)").font = f_note

    # bang doanh thu theo ngay (cot P-Q)
    ws.merge_cells("P6:Q6")
    c = ws["P6"]; c.value = "DOANH THU THEO NGÀY"
    c.font = f_head; c.fill = fill_head; c.alignment = center
    ws["P7"] = "NGÀY"; ws["Q7"] = "DOANH THU (đ)"
    for cc in ("P7", "Q7"):
        ws[cc].font = f_head; ws[cc].fill = fill_head
        ws[cc].alignment = center; ws[cc].border = border
    ndays = calendar.monthrange(YEAR, m)[1]
    for d in range(1, ndays + 1):
        rr = 7 + d
        dc = ws.cell(row=rr, column=16, value=dt.date(YEAR, m, d))
        dc.number_format = DATE_FMT; dc.border = border; dc.font = f_norm; dc.alignment = center
        vc = ws.cell(row=rr, column=17,
                     value=f"=SUMIF($B${FULL_FIRST}:$B${FULL_LAST},$P{rr},$J${FULL_FIRST}:$J${FULL_LAST})")
        vc.number_format = VND; vc.border = border; vc.font = f_norm; vc.alignment = right
    tr = 7 + ndays + 1
    tc = ws.cell(row=tr, column=16, value="CỘNG THÁNG"); tc.font = f_bold; tc.border = border
    vc = ws.cell(row=tr, column=17, value=f"=SUM($Q$8:$Q${7 + ndays})")
    vc.font = f_bold; vc.number_format = VND; vc.border = border; vc.alignment = right
    ws.freeze_panes = "A8"

for m in MONTHS:
    month_sheet(m)

# ---------------------------------------------------------------- BAO CAO NAM
ws = wb.create_sheet("BÁO CÁO NĂM")
ws.sheet_properties.tabColor = "1F4E78"
for col, w in {"A": 11, "B": 10, "C": 13, "D": 13, "E": 17,
               "F": 22, "G": 17, "H": 15, "I": 15, "J": 15}.items():
    ws.column_dimensions[col].width = w

ws.merge_cells("A1:I1")
c = ws["A1"]; c.value = f"BÁO CÁO TỔNG HỢP DOANH THU NĂM {YEAR} (THÁNG 6 → THÁNG 12)"
c.font = f_title; c.fill = fill_title; c.alignment = center
ws.row_dimensions[1].height = 26

sheets = [f"Tháng {m}" for m in MONTHS]

def sumif_all_sheets(crit_cell):
    parts = [f"SUMIF('{s}'!$B${FULL_FIRST}:$B${FULL_LAST},{crit_cell},'{s}'!$J${FULL_FIRST}:$J${FULL_LAST})"
             for s in sheets]
    return "=" + "+".join(parts)

# --- 1. Tra cuu theo ngay
ws.merge_cells("A3:D3")
c = ws["A3"]; c.value = "1. TRA CỨU DOANH THU THEO NGÀY"
c.font = f_head; c.fill = fill_head; c.alignment = left
ws.merge_cells("A4:B4")
c = ws["A4"]; c.value = "Nhập ngày cần xem:"; c.font = f_bold; c.alignment = left
ic = ws["C4"]; ic.value = dt.date(YEAR, 7, 21)
ic.number_format = DATE_FMT; ic.fill = fill_yellow; ic.font = f_blue
ic.border = border; ic.alignment = center
dc = ws["D4"]; dc.value = sumif_all_sheets("$C$4")
dc.number_format = VND; dc.font = f_bold; dc.border = border; dc.alignment = right
ws.merge_cells("A5:B5")
ws["A5"] = "→ Doanh thu ngày đó hiện ở ô D4"; ws["A5"].font = f_note

# --- 2. Theo tuan (dung theo khoi tuan nhap lieu, cot A-E)
ws.merge_cells("A7:E7")
c = ws["A7"]; c.value = "2. DOANH THU THEO TUẦN (khớp với khối tuần bạn nhập)"
c.font = f_head; c.fill = fill_head; c.alignment = left
for col, h in zip("ABCDE", ["THÁNG", "TUẦN", "TỪ NGÀY", "ĐẾN NGÀY", "DOANH THU (đ)"]):
    cc = ws[f"{col}8"]; cc.value = h
    cc.font = f_head; cc.fill = fill_head; cc.alignment = center; cc.border = border

rr = 9
for mi, m in enumerate(MONTHS):
    s = f"Tháng {m}"
    zebra = fill_sub if mi % 2 == 0 else None
    for w, (d1, d2) in enumerate(week_ranges(m), start=1):
        cells = [
            ws.cell(row=rr, column=1, value=f"Tháng {m}"),
            ws.cell(row=rr, column=2, value=f"Tuần {w}"),
            ws.cell(row=rr, column=3, value=dt.date(YEAR, m, d1)),
            ws.cell(row=rr, column=4, value=dt.date(YEAR, m, d2)),
            ws.cell(row=rr, column=5, value=(
                f"=SUMIFS('{s}'!$J${FULL_FIRST}:$J${FULL_LAST},"
                f"'{s}'!$B${FULL_FIRST}:$B${FULL_LAST},\">=\"&$C{rr},"
                f"'{s}'!$B${FULL_FIRST}:$B${FULL_LAST},\"<=\"&$D{rr})")),
        ]
        for i, cc in enumerate(cells):
            cc.font = f_norm; cc.border = border
            if zebra: cc.fill = zebra
        cells[2].number_format = DATE_FMT; cells[2].alignment = center
        cells[3].number_format = DATE_FMT; cells[3].alignment = center
        cells[4].number_format = VND; cells[4].alignment = right
        rr += 1

# --- 3. Theo thang (cot F-J)
ws.merge_cells("F7:J7")
c = ws["F7"]; c.value = "3. DOANH THU THEO THÁNG"
c.font = f_head; c.fill = fill_head; c.alignment = left
for col, h in zip("FGHIJ", ["THÁNG", "DOANH THU (đ)", "ĐÃ CK (đ)", "ĐÃ CỌC (đ)", "CHƯA CK (đ)"]):
    cc = ws[f"{col}8"]; cc.value = h
    cc.font = f_head; cc.fill = fill_head; cc.alignment = center; cc.border = border
for i, m in enumerate(MONTHS):
    r2 = 9 + i
    s = f"Tháng {m}"
    ws.cell(row=r2, column=6, value=f"{s}/{YEAR}").font = f_norm
    ws.cell(row=r2, column=6).border = border
    for col, src in ((7, "$D$3"), (8, "$D$4"), (9, "$D$5"), (10, "$D$6")):
        cc = ws.cell(row=r2, column=col, value=f"='{s}'!{src}")
        cc.number_format = VND; cc.font = f_green; cc.border = border; cc.alignment = right
m_first, m_last = 9, 9 + len(MONTHS) - 1

# --- 4. Theo quy
qrow = m_last + 2
ws.merge_cells(f"F{qrow}:J{qrow}")
c = ws[f"F{qrow}"]; c.value = "4. DOANH THU THEO QUÝ"
c.font = f_head; c.fill = fill_head; c.alignment = left
q2 = qrow + 1; q3 = qrow + 2; q4 = qrow + 3
ws.cell(row=q2, column=6, value=f"Quý 2/{YEAR} (chỉ có T6)").font = f_bold
ws.cell(row=q3, column=6, value=f"Quý 3/{YEAR} (T7-T9)").font = f_bold
ws.cell(row=q4, column=6, value=f"Quý 4/{YEAR} (T10-T12)").font = f_bold
for rq, (a, b) in ((q2, (m_first, m_first)),
                   (q3, (m_first + 1, m_first + 3)),
                   (q4, (m_first + 4, m_first + 6))):
    ws.cell(row=rq, column=6).border = border
    for col in (7, 8, 9, 10):
        L = get_column_letter(col)
        cc = ws.cell(row=rq, column=col, value=f"=SUM(${L}${a}:${L}${b})")
        cc.number_format = VND; cc.font = f_bold; cc.border = border; cc.alignment = right

# --- 5. Ca nam
yrow = q4 + 2
ws.merge_cells(f"F{yrow}:J{yrow}")
c = ws[f"F{yrow}"]; c.value = f"5. DOANH THU CẢ NĂM {YEAR}"
c.font = f_head; c.fill = fill_head; c.alignment = left
ytot = yrow + 1
ws.cell(row=ytot, column=6, value=f"TỔNG NĂM {YEAR}").font = f_bold
ws.cell(row=ytot, column=6).fill = fill_sub
ws.cell(row=ytot, column=6).border = border
for col in (7, 8, 9, 10):
    L = get_column_letter(col)
    cc = ws.cell(row=ytot, column=col, value=f"=SUM(${L}${q2}:${L}${q4})")
    cc.number_format = VND; cc.font = Font(name=FONT, size=12, bold=True, color="1F4E78")
    cc.fill = fill_sub; cc.border = border; cc.alignment = right
note = ws.cell(row=ytot + 2, column=6,
    value="Ghi chú: file theo dõi từ Tháng 6 đến Tháng 12/2026 nên Tổng năm = Quý 2 (chỉ T6) + Quý 3 + Quý 4. "
          "Tuần chia theo ngày trong tháng: 1-7, 8-14, 15-21, 22-cuối tháng.")
note.font = f_note
ws.freeze_panes = "A9"

wb.save(OUT)
print("Saved:", OUT)
print("Khoi tuan:", [block_rows(w) for w in (1, 2, 3, 4)], "| FULL_LAST =", FULL_LAST)
